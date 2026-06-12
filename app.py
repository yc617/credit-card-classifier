# -*- coding: utf-8 -*-
"""
信用卡新聞分類工具 v17
兩檔上傳版 + 爬蟲正文抽取模式：
1. Mastercard raw data
2. 信用卡分類設定與聲量總表.xlsx（含：信用卡清單_總聲量、關鍵字判定表、月份工作表）

核心原則：
- 不去重新聞；raw data 每一列都要處理。
- 同一原始列號內，同一組「銀行別 + 提及信用卡 + 卡組織」只算一次。
- 待確認、人工補卡在網頁工作臺處理；最終匯出理想上只剩「已分類」與「無卡排除」。
- v16.1 匯出改為精簡月度結果表，不再把關鍵字判定表與設定清單一併匯出。
"""

from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from io import BytesIO
from urllib.parse import urlparse
import re
import unicodedata
import warnings
import hashlib
import html as html_lib
import os
import pickle
import time

import pandas as pd
import requests
import streamlit as st
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")
try:
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
except Exception:
    pass

APP_VERSION = "v17.5.3｜Step導覽＋精簡下載區版"

MONTH_NAMES = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December"
]
MONTH_NAME_BY_NUM = {i + 1: m for i, m in enumerate(MONTH_NAMES)}
MONTH_COLUMNS = ["原始列號", "監測日期", "訊息標題", "Mastercard URL", "SourceWeb URL", "處理狀態", "銀行別", "提及信用卡", "卡組織"]

RESULT_COLUMNS = [
    "原始列號", "監測日期", "訊息標題", "Mastercard URL", "SourceWeb URL",
    "銀行別", "提及信用卡", "卡組織", "處理狀態", "判定依據"
]

PENDING_COLUMNS = RESULT_COLUMNS + ["待確認原因"]
FAILED_COLUMNS = ["原始列號", "監測日期", "訊息標題", "Mastercard URL", "SourceWeb URL", "錯誤原因"]
NO_CARD_COLUMNS = ["原始列號", "監測日期", "訊息標題", "Mastercard URL", "SourceWeb URL", "處理狀態", "判定原因"]
PAPER_SOURCE_LABEL = "報紙（監測報告無 SourceWeb）"
PAPER_REVIEW_REASON = "報紙截圖 / 監測報告無 SourceWeb，需人工審核"
TEMP_STATE_SHEET = "工作臺狀態"
TEMP_STATE_COLUMNS = MONTH_COLUMNS + ["判定依據", "待確認原因", "錯誤原因", "判定原因", "移入原因"]

SESSION_DIR = ".credit_card_classifier_sessions"
SESSION_FILE = os.path.join(SESSION_DIR, "latest_progress.pkl")
PROGRESS_KEYS = [
    "classified", "no_card_rows", "pending_rows", "failed_rows",
    "processed_orders", "no_card_orders", "pending_orders", "failed_orders",
    "last_detected", "last_pending", "last_fetch", "selected_order", "last_message",
    "manual_queue_rows", "manual_orders", "manual_active_order", "manual_staged_cards",
    "title_result_cache", "reuse_logs",
    "raw_sig", "setting_sig", "raw_bytes", "setting_bytes",
    "news_df", "card_master", "keyword_df", "org_rules", "generic_terms",
    "missing_rules_df", "orphan_rules_df",
]

REQUEST_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/121.0 Safari/537.36",
    "Accept-Language": "zh-TW,zh;q=0.9,en;q=0.8",
}

# -----------------------------------------------------------------------------
# 基礎工具
# -----------------------------------------------------------------------------


def as_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    s = str(value).strip()
    if s.lower() in {"nan", "none", "nat"}:
        return ""
    return s


def normalize_text(value) -> str:
    """NFKC + lowercase + remove whitespace for robust keyword matching."""
    s = unicodedata.normalize("NFKC", as_text(value)).lower()
    s = re.sub(r"\s+", "", s)
    return s


def normalize_key(bank: str, card: str, org: str) -> str:
    return "|".join([normalize_text(bank), normalize_text(card), normalize_text(org)])


def is_yes(value) -> bool:
    return as_text(value).upper() in {"Y", "YES", "TRUE", "1", "是", "啟用"}


def split_terms(value) -> list[str]:
    s = as_text(value)
    if not s:
        return []
    parts = re.split(r"[、,，;；|/\n\r]+", s)
    return [p.strip() for p in parts if p.strip()]


def contains_term(text: str, term: str) -> bool:
    return normalize_text(term) in normalize_text(text) if as_text(term) else False


def get_domain(url: str) -> str:
    try:
        return urlparse(as_text(url)).netloc.lower()
    except Exception:
        return ""


def is_http_url(url: str) -> bool:
    return as_text(url).lower().startswith(("http://", "https://"))


def markdown_link(label: str, url: str) -> str:
    """把 URL 轉成 Streamlit markdown 可點擊連結；非網址則保留文字。"""
    u = as_text(url)
    safe_label = html_lib.escape(as_text(label) or "開啟連結")
    if is_http_url(u):
        safe_url = html_lib.escape(u, quote=True)
        return f'<a href="{safe_url}" target="_blank" rel="noopener noreferrer">{safe_label}</a>'
    return html_lib.escape(u or "無")


def likely_mastercard_report_url(url: str) -> bool:
    domain = get_domain(url)
    return "mastercard" in domain or "rmb" in domain or "ipsos" in domain


def get_month_sheet(date_value) -> str:
    if pd.isna(date_value):
        return "January"
    try:
        dt = pd.to_datetime(date_value, errors="coerce")
        if pd.isna(dt):
            return "January"
        return MONTH_NAME_BY_NUM.get(int(dt.month), "January")
    except Exception:
        return "January"


# -----------------------------------------------------------------------------
# 讀取 Excel 設定表
# -----------------------------------------------------------------------------


def find_header_row(ws, required: list[str], max_scan: int = 10) -> int:
    required_norm = [normalize_text(x) for x in required]
    for r in range(1, min(ws.max_row, max_scan) + 1):
        row_values = [normalize_text(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if all(req in row_values for req in required_norm):
            return r
    return 1


def sheet_to_dataframe(workbook_bytes: bytes, sheet_name: str, required_headers: list[str]) -> pd.DataFrame:
    wb = load_workbook(BytesIO(workbook_bytes), data_only=True)
    if sheet_name not in wb.sheetnames:
        # fuzzy sheet match
        matched = [s for s in wb.sheetnames if sheet_name in s]
        if not matched:
            raise ValueError(f"設定表缺少工作表：{sheet_name}")
        sheet_name = matched[0]
    ws = wb[sheet_name]
    header_row = find_header_row(ws, required_headers)
    headers = [as_text(ws.cell(header_row, c).value) for c in range(1, ws.max_column + 1)]
    data = []
    for r in range(header_row + 1, ws.max_row + 1):
        row = {headers[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1) if headers[c - 1]}
        data.append(row)
    return pd.DataFrame(data)


@st.cache_data(show_spinner=False)
def load_card_master(workbook_bytes: bytes) -> pd.DataFrame:
    df = sheet_to_dataframe(workbook_bytes, "信用卡清單_總聲量", ["銀行別", "提及信用卡", "卡組織"])
    df = df.loc[:, ~df.columns.astype(str).str.contains("Unnamed")]
    rename = {"銀行": "銀行別", "卡片名稱": "提及信用卡", "正式卡名": "提及信用卡", "發卡組織": "卡組織"}
    df = df.rename(columns=rename)
    required = ["銀行別", "提及信用卡", "卡組織"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"信用卡清單_總聲量缺少欄位：{', '.join(missing)}")
    df = df[required].copy()
    df["銀行別"] = df["銀行別"].apply(as_text)
    df["提及信用卡"] = df["提及信用卡"].apply(as_text)
    df["卡組織"] = df["卡組織"].apply(as_text)
    df = df[(df["銀行別"] != "") & (df["提及信用卡"] != "")]
    return df.drop_duplicates(subset=["銀行別", "提及信用卡", "卡組織"]).reset_index(drop=True)


@st.cache_data(show_spinner=False)
def load_keyword_rules(workbook_bytes: bytes) -> pd.DataFrame:
    df = sheet_to_dataframe(workbook_bytes, "關鍵字判定表", ["主關鍵字", "提及信用卡"])
    df = df.loc[:, ~df.columns.astype(str).str.contains("Unnamed")]
    rename = {
        "關鍵字": "主關鍵字",
        "正式卡名": "提及信用卡",
        "卡片名稱": "提及信用卡",
        "發卡組織": "卡組織",
        "判定類型": "類型",
        "資料來源/備註": "資料來源",
    }
    df = df.rename(columns=rename)
    required = ["銀行別", "主關鍵字", "提及信用卡"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"關鍵字判定表缺少欄位：{', '.join(missing)}")
    defaults = {
        "類型": "精準",
        "啟用": "Y",
        "輔助關鍵字": "",
        "排除關鍵字": "",
        "卡組織": "",
        "是否通用詞": "N",
        "必須命中輔助關鍵字": "N",
        "需人工確認": "N",
        "優先級": 1,
        "判定依據": "",
        "資料來源": "",
    }
    for col, default in defaults.items():
        if col not in df.columns:
            df[col] = default
    for col in ["銀行別", "主關鍵字", "提及信用卡", "類型", "啟用", "輔助關鍵字", "排除關鍵字", "卡組織", "是否通用詞", "必須命中輔助關鍵字", "需人工確認", "判定依據", "資料來源"]:
        df[col] = df[col].apply(as_text)
    df["優先級"] = pd.to_numeric(df["優先級"], errors="coerce").fillna(2).astype(int)
    df = df[(df["主關鍵字"] != "") & (df["提及信用卡"] != "")]
    df = df[df["啟用"].apply(is_yes)].copy()
    return df.sort_values(by=["優先級"], ascending=True).reset_index(drop=True)


@st.cache_data(show_spinner=False)
def load_org_rules(workbook_bytes: bytes) -> pd.DataFrame:
    default = pd.DataFrame([
        {"判定關鍵字": "Mastercard", "卡組織": "MC"},
        {"判定關鍵字": "MasterCard", "卡組織": "MC"},
        {"判定關鍵字": "萬事達", "卡組織": "MC"},
        {"判定關鍵字": "鈦金", "卡組織": "MC"},
        {"判定關鍵字": "世界卡", "卡組織": "MC"},
        {"判定關鍵字": "Visa", "卡組織": "VISA"},
        {"判定關鍵字": "VISA", "卡組織": "VISA"},
        {"判定關鍵字": "御璽", "卡組織": "VISA"},
        {"判定關鍵字": "無限", "卡組織": "VISA"},
        {"判定關鍵字": "JCB", "卡組織": "JCB"},
        {"判定關鍵字": "晶緻", "卡組織": "JCB"},
        {"判定關鍵字": "極緻", "卡組織": "JCB"},
    ])
    try:
        df = sheet_to_dataframe(workbook_bytes, "卡組織判定表", ["判定關鍵字", "卡組織"])
        if "啟用" in df.columns:
            df = df[df["啟用"].apply(is_yes)]
        df["判定關鍵字"] = df["判定關鍵字"].apply(as_text)
        df["卡組織"] = df["卡組織"].apply(as_text)
        df = df[(df["判定關鍵字"] != "") & (df["卡組織"] != "")]
        return df[["判定關鍵字", "卡組織"]].reset_index(drop=True) if len(df) else default
    except Exception:
        return default


@st.cache_data(show_spinner=False)
def load_generic_terms(workbook_bytes: bytes) -> set[str]:
    terms = set()
    try:
        df = sheet_to_dataframe(workbook_bytes, "通用詞與銀行字典", ["通用詞"])
        if "通用詞" in df.columns:
            for value in df["通用詞"].dropna().tolist():
                t = as_text(value)
                if t:
                    terms.add(normalize_text(t))
    except Exception:
        pass
    defaults = ["現金回饋卡", "現金回饋", "商旅卡", "商務卡", "一卡通聯名卡", "聯名卡", "世界卡", "無限卡", "晶緻卡", "御璽卡", "鈦金卡", "企業卡", "銀行卡", "指定信用卡", "指定卡別", "卡友", "刷卡"]
    terms.update(normalize_text(x) for x in defaults)
    return terms


# -----------------------------------------------------------------------------
# 讀 raw data
# -----------------------------------------------------------------------------


def read_raw_news(uploaded_file) -> pd.DataFrame:
    """讀取 Mastercard raw data。優先讀取標題 cell hyperlink；不做新聞去重。"""
    name = getattr(uploaded_file, "name", "").lower()
    rows = []
    if name.endswith(".xlsx"):
        uploaded_file.seek(0)
        wb = load_workbook(uploaded_file, data_only=True)
        ws = wb.active
        headers = [as_text(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
        # 找欄位；預設 A 日期、B 標題。
        date_col = next((i + 1 for i, h in enumerate(headers) if "監測日期" in h or h == "日期"), 1)
        title_col = next((i + 1 for i, h in enumerate(headers) if "訊息標題" in h or "標題" in h), 2)
        url_col = next((i + 1 for i, h in enumerate(headers) if h in {"網址", "URL", "url"} or "網址" in h), None)
        for excel_row in range(2, ws.max_row + 1):
            date_cell = ws.cell(excel_row, date_col)
            title_cell = ws.cell(excel_row, title_col)
            title = as_text(title_cell.value)
            if not title:
                continue
            report_url = ""
            if title_cell.hyperlink:
                report_url = as_text(title_cell.hyperlink.target)
            if not report_url and url_col:
                url_cell = ws.cell(excel_row, url_col)
                report_url = as_text(url_cell.value)
                if url_cell.hyperlink:
                    report_url = as_text(url_cell.hyperlink.target) or report_url
            rows.append({
                "原始列號": len(rows) + 1,
                "監測日期": date_cell.value,
                "訊息標題": title,
                "Mastercard URL": report_url,
            })
    else:
        uploaded_file.seek(0)
        df = pd.read_excel(uploaded_file)
        df.columns = [as_text(c) for c in df.columns]
        date_col = next((c for c in df.columns if "監測日期" in c or c == "日期"), df.columns[0])
        title_col = next((c for c in df.columns if "訊息標題" in c or "標題" in c), df.columns[1] if len(df.columns) > 1 else df.columns[0])
        url_col = next((c for c in df.columns if "網址" in c or c.lower() == "url"), None)
        for _, row in df.iterrows():
            title = as_text(row.get(title_col, ""))
            if not title:
                continue
            rows.append({
                "原始列號": len(rows) + 1,
                "監測日期": row.get(date_col, ""),
                "訊息標題": title,
                "Mastercard URL": as_text(row.get(url_col, "")) if url_col else "",
            })
    return pd.DataFrame(rows)


def file_signature(file_name: str, file_bytes: bytes) -> str:
    return f"{as_text(file_name)}|{len(file_bytes)}|{hashlib.md5(file_bytes).hexdigest()}"


@st.cache_data(show_spinner=False)
def read_raw_news_cached(file_bytes: bytes, file_name: str) -> pd.DataFrame:
    bio = BytesIO(file_bytes)
    bio.name = file_name
    return read_raw_news(bio)


def reset_working_state(reason: str = ""):
    """上傳檔案更換時清空工作臺結果，避免沿用上一份資料。"""
    for key in [
        "classified", "no_card_rows", "pending_rows", "failed_rows",
        "processed_orders", "no_card_orders", "pending_orders", "failed_orders",
        "last_detected", "last_pending", "last_fetch", "manual_queue_rows",
        "manual_orders", "manual_active_order", "manual_staged_cards", "title_result_cache", "reuse_logs",
        "temp_workbook_bytes", "final_workbook_bytes", "download_status_message",
    ]:
        if key in st.session_state:
            del st.session_state[key]
    init_state()
    if reason:
        st.session_state.last_message = reason


def _ensure_session_dir():
    try:
        os.makedirs(SESSION_DIR, exist_ok=True)
    except Exception:
        pass


def save_progress_to_disk(reason: str = "auto") -> bool:
    """把目前工作臺進度存到本機/部署環境的暫存檔，降低重刷後遺失風險。"""
    try:
        if not isinstance(st.session_state.get("news_df"), pd.DataFrame) or st.session_state.news_df.empty:
            return False
        _ensure_session_dir()
        payload = {"saved_at": time.time(), "reason": reason, "app_version": APP_VERSION, "data": {}}
        for key in PROGRESS_KEYS:
            if key in st.session_state:
                payload["data"][key] = st.session_state[key]
        with open(SESSION_FILE, "wb") as f:
            pickle.dump(payload, f)
        st.session_state["last_autosave_at"] = payload["saved_at"]
        return True
    except Exception as exc:
        st.session_state["last_autosave_error"] = str(exc)
        return False


def load_saved_progress() -> dict | None:
    try:
        if not os.path.exists(SESSION_FILE):
            return None
        with open(SESSION_FILE, "rb") as f:
            return pickle.load(f)
    except Exception:
        return None


def restore_progress_to_state(payload: dict) -> bool:
    try:
        data = payload.get("data", {}) if isinstance(payload, dict) else {}
        for key, value in data.items():
            st.session_state[key] = value
        st.session_state["progress_restored"] = True
        st.session_state["last_autosave_at"] = payload.get("saved_at", None)
        return True
    except Exception as exc:
        st.session_state["last_autosave_error"] = str(exc)
        return False


def clear_saved_progress_file() -> bool:
    try:
        if os.path.exists(SESSION_FILE):
            os.remove(SESSION_FILE)
        return True
    except Exception:
        return False


def saved_progress_label(payload: dict | None) -> str:
    if not payload:
        return "目前沒有可恢復的上次進度。"
    saved_at = payload.get("saved_at")
    try:
        ts = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(saved_at)) if saved_at else "未知時間"
    except Exception:
        ts = "未知時間"
    data = payload.get("data", {})
    classified_count = len(data.get("classified", [])) if data.get("classified") is not None else 0
    pending_count = len(data.get("pending_rows", [])) if data.get("pending_rows") is not None else 0
    no_card_count = len(data.get("no_card_rows", [])) if data.get("no_card_rows") is not None else 0
    return f"上次自動保存：{ts}｜已分類 {classified_count}｜待確認 {pending_count}｜無卡排除 {no_card_count}"


def _ensure_df_columns(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    """補齊 DataFrame 欄位，避免由暫存 Excel 恢復時缺欄造成後續表格錯誤。"""
    if df is None or not isinstance(df, pd.DataFrame):
        df = pd.DataFrame()
    out = df.copy()
    for col in columns:
        if col not in out.columns:
            out[col] = ""
    return out[columns].copy()


def _row_has_value(row: dict) -> bool:
    return any(as_text(v) for v in row.values())


def _read_table_sheet(wb, sheet_name: str) -> list[dict]:
    """讀取第一列為表頭的工作表。"""
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    headers = [as_text(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        item = {}
        for c, h in enumerate(headers, start=1):
            if h:
                item[h] = ws.cell(r, c).value
        if _row_has_value(item):
            rows.append(item)
    return rows


def read_temp_workbook_progress(workbook_bytes: bytes) -> dict:
    """讀取「暫存月度結果表」，重建工作臺狀態。

    支援新版含「工作臺狀態」分頁，也支援舊版只含 January-December 月份明細的暫存表。
    """
    wb = load_workbook(BytesIO(workbook_bytes), data_only=True)
    rows = _read_table_sheet(wb, TEMP_STATE_SHEET)
    if not rows:
        # 兼容舊版暫存表：直接從月份明細表恢復。
        for sheet in MONTH_NAMES:
            rows.extend(_read_table_sheet(wb, sheet))
    if not rows:
        raise ValueError("暫存結果表中找不到可恢復的月份明細或工作臺狀態。")

    # 補齊欄位，並排除沒有原始列號的列。
    clean_rows = []
    for raw in rows:
        item = {c: raw.get(c, "") for c in TEMP_STATE_COLUMNS}
        try:
            item["原始列號"] = int(float(as_text(item.get("原始列號", ""))))
        except Exception:
            continue
        item["處理狀態"] = as_text(item.get("處理狀態", "")) or "未處理"
        clean_rows.append(item)
    if not clean_rows:
        raise ValueError("暫存結果表沒有可辨識的原始列號。")

    # 由所有狀態列重建 raw data 索引。
    base_map: dict[int, dict] = {}
    for r in clean_rows:
        order = int(r["原始列號"])
        if order not in base_map:
            base_map[order] = {
                "原始列號": order,
                "監測日期": r.get("監測日期", ""),
                "訊息標題": r.get("訊息標題", ""),
                "Mastercard URL": r.get("Mastercard URL", ""),
            }
        else:
            # 若第一筆缺資料，用後續列補上。
            for k in ["監測日期", "訊息標題", "Mastercard URL"]:
                if not as_text(base_map[order].get(k, "")) and as_text(r.get(k, "")):
                    base_map[order][k] = r.get(k, "")
    news_df = pd.DataFrame([base_map[k] for k in sorted(base_map)])

    classified, no_card_rows, pending_rows, failed_rows, manual_queue_rows = [], [], [], [], []
    for r in clean_rows:
        status = normalize_text(r.get("處理狀態", ""))
        if status == normalize_text("已分類") and as_text(r.get("銀行別", "")) and as_text(r.get("提及信用卡", "")):
            classified.append({
                "原始列號": r.get("原始列號", ""),
                "監測日期": r.get("監測日期", ""),
                "訊息標題": r.get("訊息標題", ""),
                "Mastercard URL": r.get("Mastercard URL", ""),
                "SourceWeb URL": r.get("SourceWeb URL", ""),
                "銀行別": r.get("銀行別", ""),
                "提及信用卡": r.get("提及信用卡", ""),
                "卡組織": r.get("卡組織", ""),
                "處理狀態": "已分類",
                "判定依據": r.get("判定依據", "") or "由暫存月度結果表恢復",
            })
        elif status == normalize_text("無卡排除"):
            no_card_rows.append({
                "原始列號": r.get("原始列號", ""),
                "監測日期": r.get("監測日期", ""),
                "訊息標題": r.get("訊息標題", ""),
                "Mastercard URL": r.get("Mastercard URL", ""),
                "SourceWeb URL": r.get("SourceWeb URL", ""),
                "處理狀態": "無卡排除",
                "判定原因": r.get("判定原因", "") or "由暫存月度結果表恢復",
            })
        elif status == normalize_text("待確認"):
            pending_rows.append({
                "原始列號": r.get("原始列號", ""),
                "監測日期": r.get("監測日期", ""),
                "訊息標題": r.get("訊息標題", ""),
                "Mastercard URL": r.get("Mastercard URL", ""),
                "SourceWeb URL": r.get("SourceWeb URL", ""),
                "銀行別": r.get("銀行別", ""),
                "提及信用卡": r.get("提及信用卡", ""),
                "卡組織": r.get("卡組織", ""),
                "處理狀態": "待確認",
                "判定依據": r.get("判定依據", "") or "由暫存月度結果表恢復",
                "待確認原因": r.get("待確認原因", "") or "由暫存月度結果表恢復",
            })
        elif status == normalize_text("抓取失敗") or "失敗" in status:
            failed_rows.append({
                "原始列號": r.get("原始列號", ""),
                "監測日期": r.get("監測日期", ""),
                "訊息標題": r.get("訊息標題", ""),
                "Mastercard URL": r.get("Mastercard URL", ""),
                "SourceWeb URL": r.get("SourceWeb URL", ""),
                "錯誤原因": r.get("錯誤原因", "") or "由暫存月度結果表恢復",
            })
        elif status == normalize_text("手動補卡"):
            manual_queue_rows.append({
                "原始列號": r.get("原始列號", ""),
                "監測日期": r.get("監測日期", ""),
                "訊息標題": r.get("訊息標題", ""),
                "Mastercard URL": r.get("Mastercard URL", ""),
                "SourceWeb URL": r.get("SourceWeb URL", ""),
                "移入原因": r.get("移入原因", "") or "由暫存月度結果表恢復",
                "全文": "",
                "內文標註": "",
            })
        # 未處理只需要留在 news_df，不加入任何工作臺集合。

    classified_df = _ensure_df_columns(pd.DataFrame(classified), RESULT_COLUMNS)
    return {
        "news_df": news_df,
        "classified": classified_df,
        "no_card_rows": no_card_rows,
        "pending_rows": pending_rows,
        "failed_rows": failed_rows,
        "manual_queue_rows": manual_queue_rows,
    }


def rebuild_title_result_cache_from_state():
    """從已恢復的已分類 / 無卡排除結果重建同標題快取。"""
    st.session_state.title_result_cache = {}
    if not isinstance(st.session_state.get("news_df"), pd.DataFrame) or st.session_state.news_df.empty:
        return
    for _, row in st.session_state.news_df.iterrows():
        try:
            order = int(row.get("原始列號"))
        except Exception:
            continue
        key = reusable_title_key(row.get("訊息標題", ""))
        if not key or key in st.session_state.title_result_cache:
            continue
        classified = []
        if isinstance(st.session_state.get("classified"), pd.DataFrame) and len(st.session_state.classified):
            classified = st.session_state.classified[st.session_state.classified["原始列號"].astype(str) == str(order)].to_dict("records")
        no_cards = [r for r in st.session_state.get("no_card_rows", []) if str(r.get("原始列號", "")) == str(order)]
        if classified:
            st.session_state.title_result_cache[key] = {"source_order": order, "type": "已分類", "classified": classified, "no_card": None}
        elif no_cards:
            st.session_state.title_result_cache[key] = {"source_order": order, "type": "無卡排除", "classified": [], "no_card": no_cards[0]}


def restore_temp_workbook_progress_to_state(progress: dict):
    """將暫存月度結果表的狀態寫回 session_state，讓後續可繼續偵測與人工審核。"""
    st.session_state.news_df = progress.get("news_df", pd.DataFrame())
    st.session_state.classified = _ensure_df_columns(progress.get("classified", pd.DataFrame()), RESULT_COLUMNS)
    st.session_state.no_card_rows = progress.get("no_card_rows", []) or []
    st.session_state.pending_rows = progress.get("pending_rows", []) or []
    st.session_state.failed_rows = progress.get("failed_rows", []) or []
    st.session_state.manual_queue_rows = progress.get("manual_queue_rows", []) or []

    st.session_state.processed_orders = set(st.session_state.classified["原始列號"].astype(int).tolist()) if len(st.session_state.classified) else set()
    st.session_state.no_card_orders = set(int(r.get("原始列號")) for r in st.session_state.no_card_rows if as_text(r.get("原始列號", "")))
    st.session_state.pending_orders = set(int(r.get("原始列號")) for r in st.session_state.pending_rows if as_text(r.get("原始列號", "")))
    st.session_state.failed_orders = set(int(r.get("原始列號")) for r in st.session_state.failed_rows if as_text(r.get("原始列號", "")))
    st.session_state.manual_orders = set(int(r.get("原始列號")) for r in st.session_state.manual_queue_rows if as_text(r.get("原始列號", "")))

    st.session_state.last_detected = pd.DataFrame()
    st.session_state.last_pending = pd.DataFrame()
    st.session_state.last_fetch = {}
    st.session_state.selected_order = int(st.session_state.news_df.iloc[0]["原始列號"]) if len(st.session_state.news_df) else None
    st.session_state.manual_active_order = next(iter(sorted(st.session_state.manual_orders)), None) if st.session_state.manual_orders else None
    st.session_state.manual_staged_cards = {}
    st.session_state.reuse_logs = []
    rebuild_title_result_cache_from_state()
    sort_state_tables()
    invalidate_downloads()


# -----------------------------------------------------------------------------
# 網頁抓文
# -----------------------------------------------------------------------------


@st.cache_data(show_spinner=False, ttl=3600)
def fetch_html(url: str, timeout: int = 15) -> tuple[str, str | None]:
    if not is_http_url(url):
        return "", "沒有可用網址"
    try:
        resp = requests.get(url, headers=REQUEST_HEADERS, timeout=timeout, verify=False)
        resp.raise_for_status()
        # requests usually detects encoding; force apparent when missing.
        if not resp.encoding or resp.encoding.lower() == "iso-8859-1":
            resp.encoding = resp.apparent_encoding
        return resp.text, None
    except Exception as e:
        return "", str(e)


def extract_source_url_from_report_html(html: str) -> str:
    soup = BeautifulSoup(html or "", "lxml")
    links = []
    for a in soup.find_all("a", href=True):
        href = as_text(a.get("href"))
        if not href.startswith(("http://", "https://")):
            continue
        domain = get_domain(href)
        if not domain:
            continue
        if any(x in domain for x in ["mastercard", "rmb", "ipsos"]):
            continue
        text = as_text(a.get_text(" "))
        score = 0
        near = as_text(a.parent.get_text(" ") if a.parent else "")
        if any(x.lower() in (text + near).lower() for x in ["sourceweb", "原文", "原始", "網址", "新聞"]):
            score += 10
        links.append((score, href))
    if not links:
        return ""
    links.sort(key=lambda x: x[0], reverse=True)
    return links[0][1]



NON_ARTICLE_LINE_PATTERNS = [
    "廣告", "熱門新聞", "相關新聞", "你可能也喜歡", "你可能有興趣", "更多新聞",
    "推薦閱讀", "熱門文章", "下一篇", "上一篇", "留言", "分享", "加入 Google 偏好來源",
    "訂閱", "看更多", "相關連結", "延伸閱讀", "開卡文》", "信用卡推薦",
]


def clean_line_for_article(line: str) -> str:
    line = re.sub(r"\s+", " ", as_text(line)).strip()
    line = re.sub(r"^[\s\u3000\-–—•●▲★☆▶👉]+", "", line).strip()
    return line


def is_credit_card_signal_line(line: str) -> bool:
    """新聞後段若出現信用卡訊號，即使段落較短也應保留給偵測。"""
    terms = [
        "信用卡", "聯名卡", "卡友", "刷卡", "辦卡", "首刷", "持卡", "卡片",
        "亞洲萬里通", "Asia Miles", "Mastercard", "Visa", "VISA", "JCB",
        "現金回饋", "哩程", "哩數", "紅利", "小樹點", "OPEN POINT"
    ]
    return any(contains_term(line, t) for t in terms)


def is_probable_article_paragraph(line: str) -> bool:
    """判斷段落是否像新聞主文，而不是推薦標題 / 廣告 / UI 文字。"""
    line = clean_line_for_article(line)
    if not line:
        return False
    norm = normalize_text(line)
    if len(norm) <= 8:
        return False
    if any(contains_term(line, x) for x in ["加入 Google 偏好來源", "0", "廣告"]):
        return False
    if any(contains_term(line, x) for x in ["熱門新聞", "相關新聞", "你可能也喜歡", "你可能有興趣", "下一篇", "上一篇", "開卡文》", "信用卡推薦"]):
        return False
    # 短句且像標題/連結，不當作正文；但若本身含信用卡訊號，保留給後段正文偵測。
    if len(line) < 28 and not re.search(r"[。！？；，、]", line) and not is_credit_card_signal_line(line):
        return False
    return True


def strip_bad_html_nodes(soup: BeautifulSoup, site: str = "") -> BeautifulSoup:
    for tag in soup(["script", "style", "noscript", "svg", "form", "iframe", "nav", "footer", "header", "aside"]):
        tag.decompose()
    bad_tokens = [
        "ad", "advert", "promo", "recommend", "related", "sidebar", "cookie", "share",
        "popular", "comment", "footer", "header", "nav", "social", "subscribe", "newsletter",
    ]
    for tag in soup.find_all(True):
        try:
            raw_attrs = getattr(tag, "attrs", None)
            if not isinstance(raw_attrs, dict):
                continue
            tag_id = as_text(raw_attrs.get("id", ""))
            tag_classes = raw_attrs.get("class", [])
            if isinstance(tag_classes, (list, tuple, set)):
                class_text = " ".join(as_text(x) for x in tag_classes)
            else:
                class_text = as_text(tag_classes)
            attrs = " ".join([tag_id, class_text]).lower()
        except Exception:
            continue
        if any(x in attrs for x in bad_tokens):
            # Yahoo 的正文容器可能有 content/body 字樣；不要過度刪除主要內容。
            if any(keep in attrs for keep in ["article", "story", "content", "body", "caas"]):
                continue
            try:
                tag.decompose()
            except Exception:
                pass
    return soup


def extract_paragraphs_from_nodes(nodes) -> list[str]:
    paragraphs = []
    seen = set()
    for node in nodes:
        # 優先逐段取 p/li，避免整個容器把推薦區混進來。
        children = node.find_all(["p", "li", "h1", "h2", "h3"], recursive=True) if hasattr(node, "find_all") else []
        if not children:
            children = [node]
        for child in children:
            text = clean_line_for_article(child.get_text(" ", strip=True) if hasattr(child, "get_text") else as_text(child))
            if not text:
                continue
            key = normalize_text(text)
            if key in seen:
                continue
            seen.add(key)
            paragraphs.append(text)
    return paragraphs


def extract_yahoo_article_text(soup: BeautifulSoup, url: str = "") -> str:
    """Yahoo 新聞正文抽取。

    重點：延伸閱讀只視為要跳過的區塊，不作為全文終止點；後方若仍是長段落正文，仍保留。
    """
    soup = strip_bad_html_nodes(soup, site="yahoo")
    selectors = [
        "article", "main", "div.caas-body", "div[data-test-locator='articleBody']",
        "div[class*='caas-body']", "div[class*='article']", "div[class*='story']",
    ]
    nodes = []
    for sel in selectors:
        nodes.extend(soup.select(sel))
    if not nodes and soup.body:
        nodes = [soup.body]
    paragraphs = extract_paragraphs_from_nodes(nodes)

    kept = []
    skip_related_count = 0
    for line in paragraphs:
        # 看到延伸閱讀/相關新聞時，不 break，只跳過其後幾則短標題；遇到長正文會恢復保留。
        if any(contains_term(line, x) for x in ["延伸閱讀", "相關新聞", "熱門新聞", "你可能也喜歡", "你可能有興趣"]):
            skip_related_count = 4
            continue
        if skip_related_count > 0:
            # Yahoo 的「延伸閱讀」後面有時會回到新聞主文。只跳過短連結標題；
            # 若遇到長段落或信用卡訊號段落，恢復保留。
            if (is_probable_article_paragraph(line) and len(line) >= 32) or is_credit_card_signal_line(line):
                skip_related_count = 0
            else:
                skip_related_count -= 1
                continue
        if is_probable_article_paragraph(line):
            kept.append(line)
    if not kept:
        kept = [x for x in paragraphs if is_probable_article_paragraph(x)]
    return "\n".join(kept).strip()


def is_related_link_like(line: str) -> bool:
    """判斷是否像延伸閱讀/相關新聞連結標題。"""
    line = clean_line_for_article(line)
    if not line:
        return True
    if any(contains_term(line, x) for x in ["延伸閱讀", "相關新聞", "熱門新聞", "你可能也喜歡", "你可能有興趣", "下一篇", "上一篇", "Post navigation"]):
        return True
    # 短句、沒有句讀、像標題連結時跳過；但不要把信用卡主文長段落排掉。
    if len(line) < 42 and not re.search(r"[。！？；，]", line):
        return True
    return False


def extract_udn_article_text(soup: BeautifulSoup, url: str = "") -> str:
    """經濟日報 / UDN 正文抽取。

    規則：
    - 延伸閱讀只跳過延伸閱讀連結區，不把其標題當正文。
    - 經濟日報正文遇到「※ 歡迎用『轉貼』或『分享』...」版權提示後即停止。
    """
    soup = strip_bad_html_nodes(soup, site="udn")
    selectors = [
        "article", "main", "#article_body", "#story_body_content", ".article-content",
        ".article_body", ".story_body_content", ".story-content", ".article__content",
        "div[class*='article']", "div[class*='story']", "div[class*='content']",
    ]
    nodes = []
    for sel in selectors:
        nodes.extend(soup.select(sel))
    if not nodes and soup.body:
        nodes = [soup.body]
    paragraphs = extract_paragraphs_from_nodes(nodes)
    kept = []
    skip_related_count = 0
    for line in paragraphs:
        if contains_term(line, "※ 歡迎用") or contains_term(line, "未經授權，請勿複製轉貼文章內容"):
            break
        if any(contains_term(line, x) for x in ["延伸閱讀", "相關新聞", "熱門新聞", "你可能也想看", "看更多", "推薦閱讀"]):
            skip_related_count = 5
            continue
        if skip_related_count > 0:
            if is_related_link_like(line):
                skip_related_count -= 1
                continue
            skip_related_count = 0
        if is_probable_article_paragraph(line):
            kept.append(line)
    if not kept:
        kept = [x for x in paragraphs if is_probable_article_paragraph(x)]
    return "\n".join(dict.fromkeys(kept)).strip()


def extract_technews_article_text(soup: BeautifulSoup, url: str = "") -> str:
    """TechNews 正文抽取。

    TechNews 的延伸閱讀、請喝咖啡、Post navigation 與側欄很多，
    正文通常在延伸閱讀前結束，因此遇到延伸閱讀後停止。
    """
    soup = strip_bad_html_nodes(soup, site="technews")
    selectors = ["article", "main", ".indent", ".article-content", ".entry-content", "div[class*='content']"]
    nodes = []
    for sel in selectors:
        nodes.extend(soup.select(sel))
    if not nodes and soup.body:
        nodes = [soup.body]
    paragraphs = extract_paragraphs_from_nodes(nodes)
    kept = []
    for line in paragraphs:
        if any(contains_term(line, x) for x in ["延伸閱讀", "文章看完覺得有幫助", "請我們喝杯咖啡", "Post navigation", "本週熱門", "編輯精選", "財訊快報"]):
            if len("\n".join(kept)) > 250:
                break
            continue
        if is_probable_article_paragraph(line) or (is_credit_card_signal_line(line) and len(line) >= 14):
            kept.append(line)
    if not kept:
        kept = [x for x in paragraphs if is_probable_article_paragraph(x)]
    return "\n".join(dict.fromkeys(kept)).strip()


def extract_cardu_article_text(soup: BeautifulSoup, url: str = "") -> str:
    """CardU 正文抽取：保留主內容、排除頁首廣告/推薦/下一篇。"""
    soup = strip_bad_html_nodes(soup, site="cardu")
    selectors = ["article", "main", "div[class*='article']", "div[class*='content']", "div[class*='news']", "#content"]
    nodes = []
    for sel in selectors:
        nodes.extend(soup.select(sel))
    if not nodes and soup.body:
        nodes = [soup.body]
    paragraphs = extract_paragraphs_from_nodes(nodes)
    kept = []
    for line in paragraphs:
        # CardU 的下一篇/推薦區通常應終止；但若前面正文太少，先略過該行，避免整篇變成全文過短。
        if any(contains_term(line, x) for x in ["熱門新聞", "相關新聞", "你可能有興趣", "下一篇", "上一篇"]):
            if len("\n".join(kept)) > 300:
                break
            continue
        if is_cardu_noise_line(line):
            continue
        kept.append(line)
    text = "\n".join(kept).strip()
    # 如果抽文過短，改用較寬鬆正文 fallback，避免 CardU 被誤判抓取全文過短而完全無法偵測。
    if len(text) < 250 and soup.body:
        loose = []
        for raw in soup.body.get_text("\n", strip=True).splitlines():
            line = clean_line_for_article(raw)
            if not line:
                continue
            if any(contains_term(line, x) for x in ["熱門新聞", "相關新聞", "下一篇", "上一篇", "你可能有興趣"]):
                if len("\n".join(loose)) > 300:
                    break
                continue
            if is_cardu_noise_line(line):
                continue
            if is_probable_article_paragraph(line) or is_credit_card_signal_line(line) or len(line) >= 18:
                loose.append(line)
        loose_text = "\n".join(dict.fromkeys(loose)).strip()
        if len(loose_text) > len(text):
            text = loose_text
    return text


def extract_generic_article_text(soup: BeautifulSoup, url: str = "") -> str:
    soup = strip_bad_html_nodes(soup, site="generic")
    candidates = []
    for selector in ["article", "main", ".article", ".article-content", ".news-content", ".story-body", "#article", "#content"]:
        for node in soup.select(selector):
            paragraphs = extract_paragraphs_from_nodes([node])
            text = "\n".join([x for x in paragraphs if is_probable_article_paragraph(x)]).strip()
            if len(text) > 120:
                candidates.append(text)
    if not candidates and soup.body:
        paragraphs = extract_paragraphs_from_nodes([soup.body])
        text = "\n".join([x for x in paragraphs if is_probable_article_paragraph(x)]).strip()
        if text:
            candidates.append(text)
    text = max(candidates, key=len) if candidates else soup.get_text("\n", strip=True)
    return text.strip()


def extract_text_from_html(html: str, url: str = "") -> str:
    """站台式正文抽取入口。

    v16 crawler extraction mode：
    - Yahoo：延伸閱讀只跳過區塊，不停止全文。
    - CardU：用主內容 + 後續 CardU strict detection，避免廣告/下一篇。
    - 其他站：article/main 優先，文字密度 fallback。
    """
    soup = BeautifulSoup(html or "", "lxml")
    domain = get_domain(url)
    if "yahoo" in domain:
        text = extract_yahoo_article_text(soup, url)
    elif "money.udn.com" in domain or "udn.com" in domain or "經濟日報" in as_text(soup.get_text(" ", strip=True))[:500]:
        text = extract_udn_article_text(soup, url)
    elif "technews.tw" in domain:
        text = extract_technews_article_text(soup, url)
    elif "cardu" in domain or "cardu.com" in domain:
        text = extract_cardu_article_text(soup, url)
    else:
        text = extract_generic_article_text(soup, url)
    text = re.sub(r"\n{3,}", "\n\n", text)
    text = re.sub(r"[ \t]{2,}", " ", text)
    return text.strip()


def is_paper_source_missing_error(error: str | None) -> bool:
    """監測報告沒有 SourceWeb 時，視為報紙截圖，必須進待確認。"""
    e = as_text(error)
    return "監測報告無 SourceWeb" in e or "報紙截圖" in e


def is_paper_source_label(source_url: str) -> bool:
    return "報紙" in as_text(source_url) and "SourceWeb" in as_text(source_url)

def resolve_source_and_text(input_url: str, title: str) -> tuple[str, str, str | None]:
    """回傳 source_url, article_text, error。

    抓文策略：
    1. 若是 Mastercard 監測頁，先解析 SourceWeb。
    2. 優先抓 SourceWeb 原始新聞正文。
    3. 若 SourceWeb 抓取失敗或正文過短，使用 Mastercard 監測頁文字作 fallback。
    4. 若 Mastercard 監測頁找不到 SourceWeb，判定為報紙截圖 / 掃描版，不自動無卡或已分類，後續強制進待確認。
    """
    input_url = as_text(input_url)
    if not input_url:
        note = f"【系統標註】{PAPER_REVIEW_REASON}；此列未提供可解析網址。"
        return PAPER_SOURCE_LABEL, note, PAPER_REVIEW_REASON

    source_url = input_url
    html, error = fetch_html(input_url)
    if error:
        return input_url, "", error

    if likely_mastercard_report_url(input_url):
        report_text = extract_text_from_html(html, input_url)
        extracted = extract_source_url_from_report_html(html)
        if extracted:
            source_url = extracted
            html2, error2 = fetch_html(source_url)
            if error2:
                return source_url, report_text, error2 if len(report_text) < 80 else None
            source_text = extract_text_from_html(html2, source_url)
            # SourceWeb 有抓到但正文過短時，合併 Mastercard 監測頁 fallback 文字。
            if len(as_text(source_text)) < 500 and len(as_text(report_text)) > len(as_text(source_text)):
                combined = f"{source_text}\n\n--- Mastercard fallback ---\n{report_text}".strip()
                return source_url, combined, None
            return source_url, source_text, None

        # 監測報告沒有 SourceWeb：通常代表報紙截圖 / 掃描版，不能自動無卡排除或自動分類。
        note = f"【系統標註】{PAPER_REVIEW_REASON}。此筆不進抓取失敗，也不進無卡排除；請人工開啟 Mastercard 監測報告檢查圖片或掃描內容。"
        article_text = f"{note}\n\n--- Mastercard 監測報告文字 ---\n{report_text}".strip()
        return PAPER_SOURCE_LABEL, article_text, PAPER_REVIEW_REASON

    return source_url, extract_text_from_html(html, source_url), None


def inspect_monitoring_record_without_sourceweb(input_url: str, title: str = "") -> tuple[bool, str]:
    """檢查監測紀錄是否沒有 SourceWeb。

    回傳：
    - True：監測報告沒有 SourceWeb，視為報紙截圖 / 掃描版
    - False：不是監測報告、或監測報告有 SourceWeb、或暫時無法判斷

    注意：若同標題已有「有 SourceWeb 的網路新聞」完成分類，報紙截圖可套用其卡片結果；
    若沒有可套用結果，仍必須進待確認。
    """
    input_url = as_text(input_url)
    if not input_url:
        note = f"【系統標註】{PAPER_REVIEW_REASON}；此列未提供可解析網址。"
        return True, note
    if not likely_mastercard_report_url(input_url):
        return False, ""
    html, error = fetch_html(input_url)
    if error:
        return False, ""
    extracted = extract_source_url_from_report_html(html)
    if extracted:
        return False, ""
    report_text = extract_text_from_html(html, input_url)
    note = (
        f"【系統標註】{PAPER_REVIEW_REASON}。"
        f"若同標題已有 SourceWeb 網路新聞完成分類，系統可套用其卡片分類；"
        f"若沒有，請人工開啟 Mastercard 監測報告檢查報紙圖片或掃描內容。"
    )
    article_text = f"{note}\n\n--- Mastercard 監測報告文字 ---\n{report_text}".strip()
    return True, article_text


def cached_classified_has_valid_sourceweb(cached: dict) -> bool:
    """判斷快取結果是否來自有 SourceWeb 的已分類新聞。"""
    if not cached or cached.get("type") != "已分類":
        return False
    for r in cached.get("classified", []) or []:
        url = as_text(r.get("SourceWeb URL", ""))
        if url and not is_paper_source_label(url) and "報紙" not in url:
            return True
    return False


# -----------------------------------------------------------------------------
# 偵測邏輯
# -----------------------------------------------------------------------------


def infer_org(text: str, org_rules: pd.DataFrame) -> str:
    for _, row in org_rules.iterrows():
        if contains_term(text, row.get("判定關鍵字", "")):
            return as_text(row.get("卡組織", ""))
    return ""


def canonical_org(bank: str, card: str, rule_org: str, card_master: pd.DataFrame, text: str, org_rules: pd.DataFrame) -> str:
    """優先輸出信用卡清單_總聲量中的卡組織，確保總聲量公式算得到。"""
    rule_org = as_text(rule_org)
    same = card_master[(card_master["銀行別"] == bank) & (card_master["提及信用卡"] == card)]
    if rule_org:
        matched = same[same["卡組織"].apply(as_text) == rule_org]
        if len(matched):
            return rule_org
    if len(same):
        orgs = [as_text(x) for x in same["卡組織"].tolist()]
        non_empty = [x for x in orgs if x]
        if len(non_empty) == 1:
            return non_empty[0]
        if len(orgs) and not rule_org:
            # If card master uses blank org, keep blank so formulas match.
            if "" in orgs:
                return ""
    inferred = infer_org(text, org_rules)
    if inferred:
        matched = same[same["卡組織"].apply(as_text) == inferred]
        if len(matched):
            return inferred
    return rule_org


def get_context(norm_text: str, norm_keyword: str, window: int = 120) -> str:
    idx = norm_text.find(norm_keyword)
    if idx < 0:
        return ""
    start = max(0, idx - window)
    end = min(len(norm_text), idx + len(norm_keyword) + window)
    return norm_text[start:end]


def get_match_contexts(raw_text: str, keyword: str, adjacent_lines: int = 1, fallback_window: int = 120) -> list[str]:
    """Return compact contexts that actually contain the keyword.

    This avoids matching a generic keyword in one paragraph with a bank/helper word
    that appears far away in another paragraph, which was the main source of false
    positives in CardU long-form articles.
    """
    raw_text = as_text(raw_text)
    keyword = as_text(keyword)
    if not raw_text or not keyword:
        return []
    norm_keyword = normalize_text(keyword)
    lines = [re.sub(r"\s+", " ", x).strip() for x in raw_text.splitlines()]
    lines = [x for x in lines if x]
    contexts = []
    seen = set()
    for i, line in enumerate(lines):
        if norm_keyword in normalize_text(line):
            start = max(0, i - adjacent_lines)
            end = min(len(lines), i + adjacent_lines + 1)
            ctx = "\n".join(lines[start:end]).strip()
            key = normalize_text(ctx)
            if key and key not in seen:
                seen.add(key)
                contexts.append(ctx)
    if contexts:
        return contexts

    # Fallback for text with no useful line breaks.
    norm_full = normalize_text(raw_text)
    idx = norm_full.find(norm_keyword)
    if idx < 0:
        return []
    ctx_norm = get_context(norm_full, norm_keyword, window=fallback_window)
    return [ctx_norm] if ctx_norm else []



# -----------------------------------------------------------------------------
# CardU 嚴格模式工具
# -----------------------------------------------------------------------------

CARDU_GENERIC_HEADING_EXACT = {
    normalize_text(x) for x in [
        "悠遊聯名卡", "一卡通聯名卡", "聯名卡", "信用卡", "金融卡",
        "信用白金卡", "簽帳白金卡", "白金卡", "商務卡", "商旅卡",
        "世界卡", "無限卡", "晶緻卡", "御璽卡", "鈦金卡", "JCB晶緻卡",
        "現金回饋卡", "現金回饋", "企業卡", "銀行卡"
    ]
}

SHORT_OR_AMBIGUOUS_KEYWORD_PATTERNS = [
    r"^[a-zａ-ｚA-ZＡ-Ｚ0-9０-９]{1,2}卡$",  # e卡、M卡、J卡、U卡、1卡
    r"^[a-zａ-ｚA-ZＡ-Ｚ]{1,2}\s*card$",
]

HIGH_RISK_KEYWORD_EXACT = {
    normalize_text(x) for x in [
        "e卡", "ｅ卡", "M卡", "J卡", "U卡", "一卡", "i卡",
        "悠遊聯名卡", "一卡通聯名卡", "信用白金卡", "簽帳白金卡",
        "白金卡", "世界卡", "無限卡", "晶緻卡", "御璽卡", "鈦金卡",
        "JCB晶緻卡", "Visa卡", "VISA卡", "Mastercard", "聯名卡",
    ]
}


def is_high_risk_keyword(keyword: str, rule_type: str = "", is_generic: bool = False) -> bool:
    """辨識不適合在 CardU 長文中自動分類的短詞 / 泛詞。"""
    kw = as_text(keyword)
    norm = normalize_text(kw)
    if not norm:
        return True
    if norm in HIGH_RISK_KEYWORD_EXACT:
        return True
    # v16.2：不要因為「需銀行」或「是否通用詞=Y」就一律視為高風險。
    # 使用者可能把所有卡名都改成「需銀行 + 通用詞」來避免重複規則；
    # 是否要阻擋應回到主關鍵字本身是否為短詞/泛詞，例如 e卡、世界卡、JCB晶緻卡。
    if rule_type in {"通用詞", "保護詞", "卡組織"}:
        return True
    if len(norm) <= 3 and ("卡" in norm or "card" in norm):
        return True
    for pat in SHORT_OR_AMBIGUOUS_KEYWORD_PATTERNS:
        if re.match(pat, kw, flags=re.I):
            return True
    return False


def get_cardu_header_text(block_title: str, block: str, n_lines: int = 2) -> str:
    """CardU 長文只信任區塊標題與前幾行作為主卡判斷依據。"""
    lines = [re.sub(r"\s+", " ", x).strip() for x in as_text(block).splitlines()]
    lines = [x for x in lines if x and not is_cardu_noise_line(x)]
    header_lines = []
    if as_text(block_title):
        header_lines.append(as_text(block_title))
    header_lines.extend(lines[:n_lines])
    return "\n".join(header_lines)


def cardu_rule_allowed_by_header(rule, main: str, bank: str, card: str, block_title: str, block: str, is_generic: bool, rule_type: str) -> bool:
    """CardU 嚴格模式：命中關鍵字還不夠，必須證明此區塊是在介紹該卡。"""
    if is_high_risk_keyword(main, rule_type=rule_type, is_generic=is_generic):
        return False
    header = get_cardu_header_text(block_title, block, n_lines=2)
    header_norm = normalize_text(header)
    main_norm = normalize_text(main)
    card_norm = normalize_text(card)
    bank_norm = normalize_text(bank)
    if normalize_text(block_title) in CARDU_GENERIC_HEADING_EXACT:
        return False
    if card_norm and card_norm in header_norm:
        return True
    if main_norm and main_norm in header_norm:
        return True
    if bank_norm and bank_norm in header_norm and (main_norm in header_norm or card_norm in header_norm):
        return True
    return False


def is_valid_cardu_heading_by_master(line: str, card_master: pd.DataFrame, keyword_df: pd.DataFrame) -> bool:
    """判斷 CardU 區塊標題是否像真正卡片標題，而非廣告/下一篇/泛稱。"""
    clean = re.sub(r"\s+", " ", as_text(line)).strip()
    norm = normalize_text(clean)
    if not clean or is_cardu_noise_line(clean) or len(clean) > 90:
        return False
    if norm in CARDU_GENERIC_HEADING_EXACT:
        return False
    if len(norm) <= 6 and any(x in norm for x in ["聯名卡", "白金卡", "世界卡", "晶緻卡", "御璽卡", "鈦金卡"]):
        return False
    if card_master is not None and len(card_master):
        for card in card_master.get("提及信用卡", pd.Series(dtype=str)).dropna().tolist():
            c = as_text(card)
            if c and normalize_text(c) in norm and normalize_text(c) not in CARDU_GENERIC_HEADING_EXACT:
                return True
    if keyword_df is not None and len(keyword_df):
        for _, r in keyword_df.iterrows():
            main = as_text(r.get("主關鍵字", ""))
            if not main:
                continue
            rt = as_text(r.get("類型", ""))
            gen = is_yes(r.get("是否通用詞", "N")) or rt in {"通用詞", "保護詞", "卡組織", "需銀行"}
            if is_high_risk_keyword(main, rt, gen):
                continue
            if normalize_text(main) in norm:
                return True
    markers = ["CUBE", "Unicard", "U Bear", "Richart", "DAWHO", "MaiCoin", "Gogoro", "Bankee", "iLEO", "uniopen", "Costco", "LINE Pay", "Foodpanda", "熊本熊"]
    return any(contains_term(clean, x) for x in markers) and ("卡" in clean or "card" in clean.lower())

def detect_cards_core(title: str, article_text: str, source_url: str, keyword_df: pd.DataFrame, card_master: pd.DataFrame, org_rules: pd.DataFrame, generic_terms: set[str], basis_prefix: str = "", cardu_mode: bool = False) -> tuple[pd.DataFrame, pd.DataFrame]:
    combined_text = f"{title}\n{article_text or ''}"
    detected = []
    pending = []
    seen = set()

    for _, rule in keyword_df.iterrows():
        main = as_text(rule.get("主關鍵字", ""))
        if not main:
            continue

        rule_type = as_text(rule.get("類型", "精準"))
        bank = as_text(rule.get("銀行別", ""))
        card = as_text(rule.get("提及信用卡", ""))
        if not bank or not card:
            continue

        norm_main = normalize_text(main)
        helpers = split_terms(rule.get("輔助關鍵字", ""))
        excludes = split_terms(rule.get("排除關鍵字", ""))
        is_generic_flag = is_yes(rule.get("是否通用詞", "N")) or norm_main in generic_terms
        is_rule_blocker = rule_type in {"通用詞", "保護詞", "卡組織"}
        is_hard_risky = is_high_risk_keyword(main, rule_type=rule_type, is_generic=is_generic_flag)
        is_generic = is_generic_flag or is_rule_blocker
        require_helper = is_yes(rule.get("必須命中輔助關鍵字", "N")) or rule_type in {"需銀行"} or is_generic_flag
        # v16.2：如果使用者把具體卡名標為「通用詞=Y」，只要不是短詞/泛詞且有輔助命中，仍可分類；
        # 真正的通用詞/保護詞/卡組織或短詞才進待確認/阻擋。
        review = is_yes(rule.get("需人工確認", "N")) or is_rule_blocker or (is_generic_flag and is_hard_risky) or rule_type in {"保護詞"}

        # CardU 長文嚴格模式：不要掃全文亂配；但允許「需銀行 + 非短詞」在區塊標題/前兩行成立。
        if cardu_mode:
            if is_rule_blocker or is_hard_risky:
                continue
            if not cardu_rule_allowed_by_header(rule, main, bank, card, title, article_text, is_generic_flag, rule_type):
                continue

        contexts = get_match_contexts(combined_text, main, adjacent_lines=1, fallback_window=100)
        if not contexts:
            continue

        for context in contexts:
            context_norm = normalize_text(context)
            if norm_main not in context_norm:
                continue

            # v16 短關鍵字硬性保護：e卡/ｅ卡/M卡/J卡/U卡 不能只靠整段中分散的銀行名成立。
            if re.match(r"^[a-zａ-ｚA-ZＡ-Ｚ0-9０-９]{1,2}卡$", as_text(main)):
                complete_forms = []
                if bank:
                    complete_forms.extend([f"{bank}{main}", f"{bank.replace('銀行','')}{main}", f"{bank.replace('商銀','')}{main}"])
                complete_forms.extend([f"{h}{main}" for h in helpers if h])
                if complete_forms and not any(normalize_text(x) in context_norm for x in complete_forms):
                    continue

            if excludes and any(normalize_text(x) in context_norm for x in excludes):
                continue

            helper_pool = helpers[:] if helpers else []
            # 需銀行規則若沒有填輔助關鍵字，至少要求銀行別出現在同一小段落。
            if require_helper and not helper_pool and bank:
                helper_pool = [bank]
            matched_helpers = [h for h in helper_pool if normalize_text(h) in context_norm]
            if require_helper and not matched_helpers:
                continue
            if is_rule_blocker or (is_generic_flag and is_hard_risky):
                # 真正通用詞/保護詞/卡組織/短關鍵字只作保護，不直接產生卡片結果。
                continue

            org = canonical_org(bank, card, as_text(rule.get("卡組織", "")), card_master, context, org_rules)
            key = normalize_key(bank, card, org)
            if key in seen:
                continue
            seen.add(key)
            basis = as_text(rule.get("判定依據", "")) or f"命中主關鍵字：{main}"
            if basis_prefix:
                basis = f"{basis_prefix}；" + basis
            if matched_helpers:
                basis += f"；輔助關鍵字：{'、'.join(matched_helpers)}"
            row = {
                "銀行別": bank,
                "提及信用卡": card,
                "卡組織": org,
                "處理狀態": "待確認" if review else "已分類",
                "判定依據": basis,
                "待確認原因": "規則設定需人工確認 / 保護規則" if review else "",
            }
            if review:
                pending.append(row)
            else:
                detected.append(row)
            break

    return pd.DataFrame(detected), pd.DataFrame(pending)



def is_cardu_longform_text(title: str, article_text: str, source_url: str = "") -> bool:
    domain = get_domain(source_url)
    combined = f"{title}\n{as_text(article_text)[:6000]}"
    indicators = ["推薦信用卡", "海外信用卡", "必辦", "必帶信用卡", "高回饋", "夯卡", "懶人包", "總整理", "文章目錄", "信用卡比較", "哪些銀行", "有無回饋", "刷卡有無回饋"]
    return ("cardu" in domain or "卡優" in combined) and any(contains_term(combined, x) for x in indicators)


def is_cardu_noise_line(line: str) -> bool:
    line = re.sub(r"\s+", " ", as_text(line)).strip()
    if not line:
        return True
    noise_terms = [
        "廣告", "延伸閱讀", "熱門新聞", "相關新聞", "你可能有興趣", "看更多", "更多新聞",
        "立即線上申辦", "專屬連結", "謹慎理財", "信用無價", "文章目錄",
        "開卡文》", "信用卡推薦", "下一篇", "上一篇", "Tags", "Share", "Facebook"
    ]
    if any(contains_term(line, x) for x in noise_terms):
        return True
    # CardU 頁面底部/側欄常見下一篇文章標題常有「開卡文》」等字樣；
    # 但本篇單卡標題也可能包含「永豐SPORT卡》」，不可一律排除所有「》」。
    if "》" in line and any(contains_term(line, x) for x in ["開卡文》", "信用卡推薦", "下一篇", "上一篇", "推薦"]):
        return True
    # 頁首廣告卡片常是「2026聯邦M世界卡」這種短行。
    if re.match(r"^20\d{2}.{0,12}(卡|信用卡|銀行)", line) and len(line) <= 28:
        return True
    return False


def is_probable_card_heading(line: str) -> bool:
    """保留舊接口；實際 CardU 區塊標題會在 extract_card_blocks 內嚴格驗證。"""
    line = re.sub(r"\s+", " ", as_text(line)).strip()
    if is_cardu_noise_line(line) or len(line) > 90:
        return False
    if normalize_text(line) in CARDU_GENERIC_HEADING_EXACT:
        return False
    deny = ["活動期間", "好禮", "新戶禮", "首刷禮", "了解更多", "立即申辦", "專屬連結", "信用卡推薦", "開卡文》"]
    if any(contains_term(line, x) for x in deny):
        return False
    markers = ["信用卡", "聯名卡", "認同卡", "商旅卡", "金融卡", "CUBE", "Unicard", "DAWHO", "LINE Pay", "Richart", "Gogoro", "JCB", "Bankee", "iLEO", "U Bear", "uniopen", "Panda", "利HIGH", "MaiCoin", "Costco", "熊本熊"]
    return any(contains_term(line, x) for x in markers) and (line.endswith("卡") or contains_term(line, "信用卡") or contains_term(line, "聯名卡") or contains_term(line, "金融卡"))

def clean_cardu_lines(article_text: str) -> list[str]:
    lines = [re.sub(r"\s+", " ", x).strip() for x in as_text(article_text).splitlines()]
    lines = [x for x in lines if x]
    if not lines:
        return []
    joined = "\n".join(lines)
    if "文章目錄" in joined:
        joined = joined.split("文章目錄", 1)[1]
        lines = [re.sub(r"\s+", " ", x).strip() for x in joined.splitlines() if re.sub(r"\s+", " ", x).strip()]
    cleaned = []
    for line in lines:
        # 遇到明確的頁尾/下一篇區塊，若已累積足夠正文則停止；否則略過該行繼續找正文。
        if any(contains_term(line, x) for x in ["延伸閱讀", "熱門新聞", "相關新聞", "你可能有興趣", "下一篇", "上一篇"]):
            if len("\n".join(cleaned)) > 300:
                break
            continue
        if is_cardu_noise_line(line):
            continue
        cleaned.append(line)
    return cleaned


def prepare_cardu_block_for_detection(block: str, max_lines: int = 18, max_chars: int = 1500) -> str:
    lines = [re.sub(r"\s+", " ", x).strip() for x in as_text(block).splitlines()]
    lines = [x for x in lines if x and not is_cardu_noise_line(x)]
    compact = "\n".join(lines[:max_lines]).strip()
    return compact[:max_chars]


def extract_card_blocks(article_text: str, card_master: pd.DataFrame | None = None, keyword_df: pd.DataFrame | None = None) -> list[str]:
    lines = clean_cardu_lines(article_text)
    if not lines:
        return []
    indexes = []
    seen = set()
    for idx, line in enumerate(lines):
        if is_valid_cardu_heading_by_master(line, card_master if card_master is not None else pd.DataFrame(), keyword_df if keyword_df is not None else pd.DataFrame()):
            norm = normalize_text(line)
            if norm in seen:
                continue
            seen.add(norm)
            indexes.append(idx)
    if len(indexes) < 3:
        return []
    blocks = []
    for pos, start in enumerate(indexes):
        end = indexes[pos + 1] if pos + 1 < len(indexes) else min(len(lines), start + 35)
        block = "\n".join(lines[start:end]).strip()
        block = prepare_cardu_block_for_detection(block)
        if len(block) >= 20:
            blocks.append(block)
    return blocks




# -----------------------------------------------------------------------------
# v16.5 CardU 多卡整理文 / 銀行清單 / 表格解析工具
# -----------------------------------------------------------------------------

BANK_ALIASES_FOR_LIST_PARSER = {
    "國泰世華": ["國泰世華", "國泰", "國泰世華銀行"],
    "中國信託": ["中國信託", "中信", "中國信託銀行"],
    "台新銀行": ["台新銀行", "台新"],
    "玉山銀行": ["玉山銀行", "玉山"],
    "永豐銀行": ["永豐銀行", "永豐", "永豐銀"],
    "聯邦銀行": ["聯邦銀行", "聯邦"],
    "第一銀行": ["第一銀行", "一銀", "第一銀"],
    "富邦銀行": ["富邦銀行", "台北富邦", "富邦", "北富銀"],
    "兆豐銀行": ["兆豐銀行", "兆豐", "兆豐銀"],
    "臺灣企銀": ["臺灣企銀", "台灣企銀", "台企銀", "臺企銀", "臺灣中小企銀", "台灣中小企銀"],
    "台企銀": ["臺灣企銀", "台灣企銀", "台企銀", "臺企銀", "臺灣中小企銀", "台灣中小企銀"],
    "陽信銀行": ["陽信銀行", "陽信"],
    "星展銀行": ["星展銀行", "星展"],
    "上海銀行": ["上海銀行", "上海商銀", "上海商業儲蓄銀行"],
    "遠東商銀": ["遠東商銀", "遠東銀行", "遠銀"],
    "新光銀行": ["新光銀行", "新光"],
    "彰化銀行": ["彰化銀行", "彰銀", "彰化"],
    "合作金庫": ["合作金庫", "合庫"],
    "華南銀行": ["華南銀行", "華南"],
    "元大銀行": ["元大銀行", "元大"],
    "凱基銀行": ["凱基銀行", "凱基"],
    "滙豐銀行": ["滙豐銀行", "滙豐", "匯豐", "HSBC"],
    "美國運通": ["美國運通", "Amex", "American Express"],
}

NON_CARD_LIST_TOKENS = {
    normalize_text(x) for x in [
        "全卡別", "全卡", "所有卡別", "不限卡別", "不限信用卡", "其他卡別",
        "除左側列表外全卡別", "除右側列表外全卡別", "除上述外全卡別", "除下列外全卡別",
        "無", "不適用", "不回饋", "有回饋", "回饋", "無回饋", "一般消費", "海外消費",
        "實體刷卡", "網路消費", "歐盟", "英國", "備註", "全銀行", "發卡機構",
    ]
}

MULTICARD_TITLE_INDICATORS = [
    "推薦信用卡", "信用卡推薦", "必帶信用卡", "必辦", "高回饋", "總整理", "懶人包",
    "一次看", "信用卡權益", "權益一次看", "神卡", "哩程卡", "回饋卡", "信用卡比較",
    "哪些銀行", "有無回饋", "刷卡有無回饋", "歐洲信用卡", "海外信用卡", "遊歐洲",
]

TABLE_OR_LIST_INDICATORS = [
    "發卡機構", "有回饋", "無回饋", "總整理表", "回饋總整理", "刷卡有無回饋",
    "哪些銀行", "部分卡別", "卡別", "銀行：", "銀行:", "●", "•",
]


def is_cardu_multicard_article(title: str, source_url: str, article_text: str = "") -> bool:
    if not is_cardu_site(source_url, article_text):
        return False
    combined = f"{title}\n{as_text(article_text)[:3000]}"
    return any(contains_term(combined, x) for x in MULTICARD_TITLE_INDICATORS)


def get_known_bank_aliases_from_master(card_master: pd.DataFrame | None = None) -> dict[str, list[str]]:
    aliases = dict(BANK_ALIASES_FOR_LIST_PARSER)
    if card_master is not None and len(card_master):
        for bank in card_master.get("銀行別", pd.Series(dtype=str)).dropna().astype(str).unique().tolist():
            bank = as_text(bank)
            if not bank:
                continue
            base = bank.replace("銀行", "").replace("商銀", "")
            vals = aliases.setdefault(bank, [])
            vals.extend([bank, base])
    # de-duplicate while preserving order
    out = {}
    for bank, vals in aliases.items():
        seen = set(); clean = []
        for v in vals:
            v = as_text(v)
            if not v:
                continue
            k = normalize_text(v)
            if k not in seen:
                seen.add(k); clean.append(v)
        out[bank] = clean
    return out


def find_bank_context(line: str, card_master: pd.DataFrame | None = None) -> str:
    norm = normalize_text(line)
    aliases = get_known_bank_aliases_from_master(card_master)
    # prefer longer aliases to avoid false partial matches
    candidates = []
    for bank, vals in aliases.items():
        for alias in vals:
            an = normalize_text(alias)
            if an and an in norm:
                candidates.append((len(an), bank))
    if not candidates:
        return ""
    candidates.sort(reverse=True)
    return candidates[0][1]


def is_non_card_candidate(candidate: str) -> bool:
    c = re.sub(r"\s+", " ", as_text(candidate)).strip(" ：:，,、；;。()（）[]【】")
    n = normalize_text(c)
    if not n:
        return True
    if n in NON_CARD_LIST_TOKENS:
        return True
    if any(tok in n for tok in NON_CARD_LIST_TOKENS):
        # only suppress if candidate is short enough to be mostly status text
        if len(n) <= 18:
            return True
    if len(c) <= 1:
        return True
    if c.isdigit():
        return True
    deny_terms = ["登錄", "活動", "優惠", "回饋", "現折", "最高", "每月", "上限", "限量", "詳情", "備註", "說明", "官網", "連結", "申辦", "線上"]
    if not any(x in c for x in ["卡", "card", "Card", "CUBE", "Unicard", "Richart", "DAWAY", "DAWHO", "LINE Pay", "uniopen", "MaiCoin", "Costco"]):
        return True
    if any(x in c for x in deny_terms) and len(c) <= 10:
        return True
    return False


def inherit_card_suffix(items: list[str]) -> list[str]:
    """Handle patterns like ANA 極緻卡/無限卡/晶緻卡 and 世界商務/鈦金商務卡."""
    out = []
    last_suffix = ""
    # work backwards to inherit trailing 卡 when omitted
    for item in reversed(items):
        clean = as_text(item).strip(" ：:，,、；;。()（）[]【】")
        if not clean:
            continue
        if "卡" in clean:
            last_suffix = "卡"
            out.append(clean)
        else:
            if last_suffix and any(x in clean for x in ["世界", "鈦金", "御璽", "晶緻", "商務", "無限", "白金", "極緻"]):
                out.append(clean + last_suffix)
            else:
                out.append(clean)
    return list(reversed(out))


def expand_compound_card_candidates(fragment: str) -> list[str]:
    s = as_text(fragment)
    if not s:
        return []
    # normalize brackets but keep content; remove too noisy prefixes
    s = re.sub(r"[\t\r]+", " ", s)
    s = re.sub(r"※.*$", "", s)
    s = re.sub(r"（[^）]{0,20}(?:每月|上限|回饋|需登錄|活動|歐盟|英國)[^）]{0,30}）", "", s)
    s = re.sub(r"\([^)]{0,20}(?:每月|上限|回饋|需登錄|活動|歐盟|英國)[^)]{0,30}\)", "", s)
    # split separators commonly used in CardU lists/tables
    raw_parts = re.split(r"[、,，；;\n]+|\s{2,}|(?<!https):(?=\S)", s)
    parts = []
    for part in raw_parts:
        part = as_text(part).strip()
        if not part:
            continue
        sub = re.split(r"\s*/\s*|／", part)
        expanded_sub = inherit_card_suffix(sub)
        parts.extend(expanded_sub)
    out = []
    seen = set()
    for p in parts:
        p = re.sub(r"^(有回饋|無回饋|回饋|卡別|推薦|\d+\.?\s*)", "", as_text(p)).strip(" ：:，,、；;。()（）[]【】")
        # remove bank prefix only if candidate is too long? Keep bank+card useful for matching.
        if is_non_card_candidate(p):
            continue
        k = normalize_text(p)
        if k not in seen:
            seen.add(k); out.append(p)
    return out


def extract_bank_list_candidate_texts(text: str, card_master: pd.DataFrame | None = None) -> list[dict]:
    """Parse lines like '●永豐銀行：A、B、C' or table-ish rows into focused candidate snippets."""
    lines = [re.sub(r"\s+", " ", x).strip() for x in as_text(text).splitlines()]
    lines = [x for x in lines if x and not is_cardu_noise_line(x)]
    candidates = []
    bank_aliases = get_known_bank_aliases_from_master(card_master)
    bank_alias_pattern = "|".join(sorted({re.escape(v) for vals in bank_aliases.values() for v in vals if as_text(v)}, key=len, reverse=True))
    if not bank_alias_pattern:
        return candidates
    for i, line in enumerate(lines):
        if any(contains_term(line, x) for x in ["全卡別", "除左側", "除右側", "除上述", "除下列"]):
            # still parse explicit cards on the same row, but suppress the status tokens
            pass
        bank = find_bank_context(line, card_master)
        # Pattern 1: bullet/bank colon list
        m = re.search(rf"(?:●|•|-|－)?\s*({bank_alias_pattern})\s*[:：]\s*(.+)$", line)
        if m:
            bank = find_bank_context(m.group(1), card_master) or bank
            rhs = m.group(2)
            for cand in expand_compound_card_candidates(rhs):
                candidates.append({"bank": bank, "candidate": cand, "source_line": line, "source_type": "銀行清單"})
            continue
        # Pattern 2: table-like rows where bank appears with several card names
        if bank and ("卡" in line or "CUBE" in line or "Unicard" in line or "Richart" in line):
            # remove leading bank text and status tokens; then split remaining pieces
            rhs = re.sub(rf".*?({bank_alias_pattern})", "", line, count=1).strip(" ：:｜|\t")
            # Skip pure sentence paragraphs; parse only lines that look list/table-like
            if any(sep in line for sep in ["、", "／", "/", "｜", "|", "：", ":"]) or len(line) < 80:
                for cand in expand_compound_card_candidates(rhs):
                    candidates.append({"bank": bank, "candidate": cand, "source_line": line, "source_type": "表格/銀行列"})
    # de-duplicate candidates
    out=[]; seen=set()
    for c in candidates:
        key=(normalize_text(c.get("bank","")), normalize_text(c.get("candidate","")))
        if key[1] and key not in seen:
            seen.add(key); out.append(c)
    return out


def detect_from_candidate_snippets(title: str, candidates: list[dict], source_url: str, keyword_df: pd.DataFrame, card_master: pd.DataFrame, org_rules: pd.DataFrame, generic_terms: set[str], basis_prefix: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    frames_det=[]; frames_pen=[]
    for c in candidates:
        bank = as_text(c.get("bank", ""))
        cand = as_text(c.get("candidate", ""))
        if not cand:
            continue
        # Use bank + candidate to satisfy '需銀行' helper rules without scanning unrelated page text.
        snippet = "\n".join([x for x in [bank, cand, c.get("source_line", "")] if as_text(x)])
        det, pen = detect_cards_core(title, snippet, source_url, keyword_df, card_master, org_rules, generic_terms, basis_prefix=f"{basis_prefix}/{c.get('source_type','清單')}：{cand}", cardu_mode=False)
        if len(det): frames_det.append(det)
        if len(pen): frames_pen.append(pen)
    det_all = pd.concat(frames_det, ignore_index=True) if frames_det else pd.DataFrame()
    pen_all = pd.concat(frames_pen, ignore_index=True) if frames_pen else pd.DataFrame()
    return dedupe_rule_results(det_all), dedupe_rule_results(pen_all)


def is_structured_multicard_text(title: str, article_text: str) -> bool:
    combined = f"{title}\n{as_text(article_text)[:5000]}"
    return any(contains_term(combined, x) for x in MULTICARD_TITLE_INDICATORS + TABLE_OR_LIST_INDICATORS)

def extract_list_sections(text: str) -> list[str]:
    """擷取「指定卡別：A、B、C」這類高價值區塊，避免只靠整篇全文掃描。"""
    src = as_text(text)
    if not src:
        return []
    labels = ["指定卡別", "適用卡別", "活動卡別", "指定信用卡", "適用信用卡", "申辦指定信用卡", "新戶申辦指定信用卡"]
    pattern = r"(?:" + "|".join(map(re.escape, labels)) + r")\s*[:：]\s*([^。\n]{2,350})"
    sections = []
    for m in re.finditer(pattern, src):
        sections.append(m.group(0))
    return sections




def is_cardu_site(source_url: str, article_text: str = "") -> bool:
    return "cardu" in get_domain(source_url) or "卡優" in as_text(article_text)[:300]


def is_cardu_single_card_article(title: str, source_url: str, article_text: str = "") -> bool:
    """CardU 單卡文，例如「永豐SPORT卡》...」。"""
    if not is_cardu_site(source_url, article_text):
        return False
    t = as_text(title)
    if not t or "》" not in t:
        return False
    longform_indicators = ["推薦信用卡", "信用卡推薦", "必辦", "必帶信用卡", "高回饋", "夯卡", "懶人包", "總整理", "比較", "15張", "10張", "文章目錄", "哪些銀行", "有無回饋"]
    if any(contains_term(t, x) for x in longform_indicators):
        return False
    left = t.split("》", 1)[0]
    return ("卡" in left or "card" in left.lower()) and len(left) <= 40


def filter_results_for_cardu_single(df: pd.DataFrame, title: str, keyword_df: pd.DataFrame) -> pd.DataFrame:
    """CardU 單卡文只保留標題主卡能支持的結果，避免推薦/廣告卡進入。"""
    if df is None or len(df) == 0:
        return pd.DataFrame()
    title_norm = normalize_text(title)
    allowed_keys = set()
    for _, rule in keyword_df.iterrows():
        bank = as_text(rule.get("銀行別", ""))
        card = as_text(rule.get("提及信用卡", ""))
        org = as_text(rule.get("卡組織", ""))
        main = as_text(rule.get("主關鍵字", ""))
        rt = as_text(rule.get("類型", ""))
        gen = is_yes(rule.get("是否通用詞", "N")) or rt in {"通用詞", "保護詞", "卡組織", "需銀行"}
        if not bank or not card:
            continue
        if is_high_risk_keyword(main, rt, gen):
            continue
        bank_norm = normalize_text(bank)
        card_norm = normalize_text(card)
        main_norm = normalize_text(main)
        if (card_norm and card_norm in title_norm) or (main_norm and main_norm in title_norm) or (bank_norm and bank_norm in title_norm and (card_norm in title_norm or main_norm in title_norm)):
            allowed_keys.add(normalize_key(bank, card, org))
            # 同一銀行+卡名不同卡組織也先允許，後續 canonical_org 會對齊總表。
            allowed_keys.add(normalize_key(bank, card, ""))
    if not allowed_keys:
        return pd.DataFrame()
    out = []
    for _, row in df.iterrows():
        bank = as_text(row.get("銀行別", ""))
        card = as_text(row.get("提及信用卡", ""))
        org = as_text(row.get("卡組織", ""))
        if normalize_key(bank, card, org) in allowed_keys or normalize_key(bank, card, "") in allowed_keys:
            out.append(row.to_dict())
    return pd.DataFrame(out)


def dedupe_rule_results(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or len(df) == 0:
        return pd.DataFrame()
    df = df.copy()
    df["_key"] = df.apply(lambda r: normalize_key(r.get("銀行別", ""), r.get("提及信用卡", ""), r.get("卡組織", "")), axis=1)
    # 保留先出現的高優先結果；偵測流程會把指定卡別清單放在前面。
    df = df.drop_duplicates(subset=["_key"], keep="first").drop(columns=["_key"])
    return df.reset_index(drop=True)

def detect_cards(title: str, article_text: str, source_url: str, keyword_df: pd.DataFrame, card_master: pd.DataFrame, org_rules: pd.DataFrame, generic_terms: set[str]) -> tuple[pd.DataFrame, pd.DataFrame]:
    """信用卡偵測主入口（v16.5）。

    v16.5 核心：
    - 非正文/廣告區仍嚴格排除。
    - CardU 單卡文只抓標題主卡。
    - CardU 多卡整理文改用「清單/表格/銀行列」積極解析，提高召回。
    - 一般新聞仍優先解析指定卡別清單，避免 e卡/M卡 等短詞亂入。
    """
    frames_det: list[pd.DataFrame] = []
    frames_pen: list[pd.DataFrame] = []

    cardu_site = is_cardu_site(source_url, article_text)
    cardu_multi = cardu_site and is_cardu_multicard_article(title, source_url, article_text)
    cardu_single = is_cardu_single_card_article(title, source_url, article_text) and not cardu_multi
    cardu_longform = is_cardu_longform_text(title, article_text, source_url)

    # 1) 指定卡別清單：一般新聞優先；CardU 多卡文也允許在正文內解析，因為它常包含大量正文清單。
    list_sections = extract_list_sections(f"{title}\n{article_text}") if (not cardu_site or cardu_multi) else []
    if list_sections:
        for sec in list_sections:
            det, pen = detect_cards_core(title, sec, source_url, keyword_df, card_master, org_rules, generic_terms, basis_prefix="指定卡別清單")
            if len(det): frames_det.append(det)
            if len(pen): frames_pen.append(pen)

    # 2) CardU 多卡整理文：解析「銀行：卡片清單」、表格列、斜線/頓號複合卡名。
    #    這比 strict block 更適合 10張推薦卡、歐洲回饋總整理、哪些銀行無回饋等文章。
    if cardu_multi or (cardu_site and is_structured_multicard_text(title, article_text)):
        candidates = extract_bank_list_candidate_texts(article_text, card_master)
        det_s, pen_s = detect_from_candidate_snippets(title, candidates, source_url, keyword_df, card_master, org_rules, generic_terms, basis_prefix="CardU多卡結構解析")
        if len(det_s): frames_det.append(det_s)
        if len(pen_s): frames_pen.append(pen_s)

        # 多卡文仍補跑有效卡片區塊標題，但不跑整頁全文，避免廣告區誤抓。
        blocks = extract_card_blocks(article_text, card_master, keyword_df)
        for idx, block in enumerate(blocks, start=1):
            block_title = block.splitlines()[0][:60] if block.splitlines() else f"區塊{idx}"
            det, pen = detect_cards_core(block_title, block, source_url, keyword_df, card_master, org_rules, generic_terms, basis_prefix=f"CardU多卡區塊{idx}：{block_title}", cardu_mode=True)
            if len(det): frames_det.append(det)
            if len(pen): frames_pen.append(pen)

        detected = pd.concat(frames_det, ignore_index=True) if frames_det else pd.DataFrame()
        pending = pd.concat(frames_pen, ignore_index=True) if frames_pen else pd.DataFrame()
        if len(detected) or len(pending):
            return dedupe_rule_results(detected), dedupe_rule_results(pending)
        # 如果結構解析沒抓到，才走嚴格模式，不走寬鬆全文掃描。

    # 3) 如果一般新聞指定清單已抓到結果，直接返回，避免全文短詞或頁面噪音亂入。
    if list_sections and not cardu_site:
        list_det = pd.concat(frames_det, ignore_index=True) if frames_det else pd.DataFrame()
        list_pen = pd.concat(frames_pen, ignore_index=True) if frames_pen else pd.DataFrame()
        if len(list_det) or len(list_pen):
            return dedupe_rule_results(list_det), dedupe_rule_results(list_pen)

    # 4) CardU 單卡文：只接受標題主卡支持的結果，防推薦/廣告卡。
    if cardu_single:
        det_t, pen_t = detect_cards_core(title, title, source_url, keyword_df, card_master, org_rules, generic_terms, basis_prefix="CardU單卡標題", cardu_mode=False)
        det_a, pen_a = detect_cards_core(title, article_text, source_url, keyword_df, card_master, org_rules, generic_terms, basis_prefix="CardU單卡正文", cardu_mode=False)
        det = pd.concat([det_t, det_a], ignore_index=True) if len(det_t) or len(det_a) else pd.DataFrame()
        pen = pd.concat([pen_t, pen_a], ignore_index=True) if len(pen_t) or len(pen_a) else pd.DataFrame()
        det = filter_results_for_cardu_single(det, title, keyword_df)
        pen = filter_results_for_cardu_single(pen, title, keyword_df)
        return dedupe_rule_results(det), dedupe_rule_results(pen)

    # 5) CardU 長文但非多卡結構文：使用 strict block mode。
    if cardu_site and cardu_longform:
        blocks = extract_card_blocks(article_text, card_master, keyword_df)
        if len(blocks) >= 3:
            for idx, block in enumerate(blocks, start=1):
                block_title = block.splitlines()[0][:60] if block.splitlines() else f"區塊{idx}"
                det, pen = detect_cards_core(block_title, block, source_url, keyword_df, card_master, org_rules, generic_terms, basis_prefix=f"CardU長文區塊{idx}：{block_title}", cardu_mode=True)
                if len(det): frames_det.append(det)
                if len(pen): frames_pen.append(pen)
        else:
            det, pen = detect_cards_core(title, article_text, source_url, keyword_df, card_master, org_rules, generic_terms, basis_prefix="CardU未切塊正文", cardu_mode=True)
            if len(det): frames_det.append(det)
            if len(pen): frames_pen.append(pen)
    else:
        # 6) 其他一般新聞：清理後正文 + 標題。
        det, pen = detect_cards_core(title, article_text, source_url, keyword_df, card_master, org_rules, generic_terms)
        if len(det): frames_det.append(det)
        if len(pen): frames_pen.append(pen)

    detected = pd.concat(frames_det, ignore_index=True) if frames_det else pd.DataFrame()
    pending = pd.concat(frames_pen, ignore_index=True) if frames_pen else pd.DataFrame()
    return dedupe_rule_results(detected), dedupe_rule_results(pending)

def is_title_card_related(title: str) -> bool:
    terms = ["信用卡", "刷卡", "卡友", "辦卡", "聯名卡", "御璽", "鈦金", "世界卡", "現金回饋", "首刷", "推薦信用卡", "卡優", "Richart卡", "兆豐卡", "uniopen卡", "Costco卡"]
    return any(contains_term(title, t) for t in terms)


def should_no_card(title: str, article_text: str) -> tuple[bool, str]:
    length = len(as_text(article_text))
    # 標題本身若像信用卡新聞，不因正文短就直接無卡，交給待確認/手動補卡。
    if is_title_card_related(title):
        return False, "標題疑似信用卡相關但未命中規則"
    if length < 250:
        return False, f"抓取全文字數過短：{length}"
    return True, "未命中特定信用卡"


# -----------------------------------------------------------------------------
# Session state
# -----------------------------------------------------------------------------


def init_state():
    defaults = {
        "classified": pd.DataFrame(columns=RESULT_COLUMNS),
        "no_card_rows": [],
        "pending_rows": [],
        "failed_rows": [],
        "processed_orders": set(),
        "no_card_orders": set(),
        "pending_orders": set(),
        "failed_orders": set(),
        "last_detected": pd.DataFrame(),
        "last_pending": pd.DataFrame(),
        "last_fetch": {},
        "selected_order": None,
        "last_message": "",
        "manual_queue_rows": [],
        "manual_orders": set(),
        "manual_active_order": None,
        "manual_staged_cards": {},
        "title_result_cache": {},
        "reuse_logs": [],
        "temp_workbook_bytes": None,
        "final_workbook_bytes": None,
        "download_status_message": "",
        "raw_sig": "",
        "setting_sig": "",
        "raw_bytes": b"",
        "setting_bytes": b"",
        "news_df": pd.DataFrame(),
        "card_master": pd.DataFrame(),
        "keyword_df": pd.DataFrame(),
        "org_rules": pd.DataFrame(),
        "generic_terms": set(),
        "missing_rules_df": pd.DataFrame(),
        "orphan_rules_df": pd.DataFrame(),
        "loaded_from_temp_workbook": False,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


def invalidate_downloads():
    st.session_state.temp_workbook_bytes = None
    st.session_state.final_workbook_bytes = None
    st.session_state.download_status_message = ""


def sort_state_tables():
    """讓工作臺結果依原始列號排序。
    手動補卡加入已分類後，結果會回到原始列號附近，而不是堆到最下方。
    """
    if isinstance(st.session_state.get("classified"), pd.DataFrame) and len(st.session_state.classified):
        st.session_state.classified = (
            st.session_state.classified
            .sort_values(
                by=["原始列號", "銀行別", "提及信用卡", "卡組織"],
                kind="mergesort",
                na_position="last",
            )
            .reset_index(drop=True)
        )
    for list_key in ["no_card_rows", "pending_rows", "failed_rows", "manual_queue_rows"]:
        if isinstance(st.session_state.get(list_key), list) and st.session_state[list_key]:
            st.session_state[list_key] = sorted(
                st.session_state[list_key],
                key=lambda x: (int(x.get("原始列號") or 0), as_text(x.get("銀行別", "")), as_text(x.get("提及信用卡", ""))),
            )



def remove_order(order: int):
    st.session_state.classified = st.session_state.classified[st.session_state.classified["原始列號"] != order].copy()
    st.session_state.no_card_rows = [r for r in st.session_state.no_card_rows if int(r.get("原始列號", -1)) != order]
    st.session_state.pending_rows = [r for r in st.session_state.pending_rows if int(r.get("原始列號", -1)) != order]
    st.session_state.failed_rows = [r for r in st.session_state.failed_rows if int(r.get("原始列號", -1)) != order]
    for key in ["processed_orders", "no_card_orders", "pending_orders", "failed_orders"]:
        st.session_state[key].discard(order)


def add_classified_rows(order: int, base_row: dict, detected_df: pd.DataFrame, clear_pending_order: bool = True):
    invalidate_downloads()
    order = int(order)
    # 轉為已分類時，必須清除同一原始列在其他工作區的狀態，
    # 避免「手動補卡按下後看似沒作用」或同一列同時留在無卡/失敗/手動補卡。
    st.session_state.no_card_rows = [r for r in st.session_state.no_card_rows if int(r.get("原始列號", -1)) != order]
    st.session_state.failed_rows = [r for r in st.session_state.failed_rows if int(r.get("原始列號", -1)) != order]
    st.session_state.manual_queue_rows = [r for r in st.session_state.manual_queue_rows if int(r.get("原始列號", -1)) != order]
    st.session_state.no_card_orders.discard(order)
    st.session_state.failed_orders.discard(order)
    st.session_state.manual_orders.discard(order)
    if detected_df is None or len(detected_df) == 0:
        return
    new_rows = []
    existing = set(
        normalize_key(r["銀行別"], r["提及信用卡"], r["卡組織"])
        for _, r in st.session_state.classified[st.session_state.classified["原始列號"] == order].iterrows()
    )
    for _, r in detected_df.iterrows():
        key = normalize_key(r["銀行別"], r["提及信用卡"], r["卡組織"])
        if key in existing:
            continue
        existing.add(key)
        new_rows.append({
            "原始列號": order,
            "監測日期": base_row.get("監測日期", ""),
            "訊息標題": base_row.get("訊息標題", ""),
            "Mastercard URL": base_row.get("Mastercard URL", ""),
            "SourceWeb URL": st.session_state.last_fetch.get("source_url", base_row.get("Mastercard URL", "")),
            "銀行別": r.get("銀行別", ""),
            "提及信用卡": r.get("提及信用卡", ""),
            "卡組織": r.get("卡組織", ""),
            "處理狀態": "已分類",
            "判定依據": r.get("判定依據", ""),
        })
    if new_rows:
        st.session_state.classified = pd.concat([st.session_state.classified, pd.DataFrame(new_rows)], ignore_index=True)
        sort_state_tables()
        st.session_state.processed_orders.add(order)
        if clear_pending_order:
            st.session_state.pending_orders.discard(order)
            # If user added rows, clear pending for same order to avoid unfinished block.
            st.session_state.pending_rows = [r for r in st.session_state.pending_rows if int(r.get("原始列號", -1)) != order]


def add_no_card(order: int, base_row: dict, reason: str = "人工標記無卡排除"):
    invalidate_downloads()
    remove_order(order)
    st.session_state.no_card_rows.append({
        "原始列號": order,
        "監測日期": base_row.get("監測日期", ""),
        "訊息標題": base_row.get("訊息標題", ""),
        "Mastercard URL": base_row.get("Mastercard URL", ""),
        "SourceWeb URL": st.session_state.last_fetch.get("source_url", base_row.get("Mastercard URL", "")),
        "處理狀態": "無卡排除",
        "判定原因": reason,
    })
    sort_state_tables()
    st.session_state.no_card_orders.add(order)


def add_pending(order: int, base_row: dict, pending_df: pd.DataFrame, source_url: str):
    invalidate_downloads()
    if pending_df is None or len(pending_df) == 0:
        return
    # Avoid duplicates by order + bank/card/org
    existing = set(
        normalize_key(r.get("銀行別", ""), r.get("提及信用卡", ""), r.get("卡組織", ""))
        for r in st.session_state.pending_rows if int(r.get("原始列號", -1)) == order
    )
    for _, r in pending_df.iterrows():
        key = normalize_key(r.get("銀行別", ""), r.get("提及信用卡", ""), r.get("卡組織", ""))
        if key in existing:
            continue
        existing.add(key)
        st.session_state.pending_rows.append({
            "原始列號": order,
            "監測日期": base_row.get("監測日期", ""),
            "訊息標題": base_row.get("訊息標題", ""),
            "Mastercard URL": base_row.get("Mastercard URL", ""),
            "SourceWeb URL": source_url,
            "銀行別": r.get("銀行別", ""),
            "提及信用卡": r.get("提及信用卡", ""),
            "卡組織": r.get("卡組織", ""),
            "處理狀態": "待確認",
            "判定依據": r.get("判定依據", ""),
            "待確認原因": r.get("待確認原因", ""),
        })
    sort_state_tables()
    st.session_state.pending_orders.add(order)



def pending_record_key(row: dict | pd.Series) -> str:
    return "|".join([
        as_text(row.get("原始列號", "")),
        normalize_text(row.get("銀行別", "")),
        normalize_text(row.get("提及信用卡", "")),
        normalize_text(row.get("卡組織", "")),
        normalize_text(row.get("判定依據", "")),
    ])


def refresh_pending_orders():
    orders = set()
    for r in st.session_state.pending_rows:
        try:
            orders.add(int(r.get("原始列號")))
        except Exception:
            pass
    st.session_state.pending_orders = orders


def add_failed(order: int, base_row: dict, source_url: str, error: str):
    invalidate_downloads()
    remove_order(order)
    st.session_state.failed_rows.append({
        "原始列號": order,
        "監測日期": base_row.get("監測日期", ""),
        "訊息標題": base_row.get("訊息標題", ""),
        "Mastercard URL": base_row.get("Mastercard URL", ""),
        "SourceWeb URL": source_url,
        "錯誤原因": error,
    })
    st.session_state.failed_orders.add(order)


def status_for(order: int) -> str:
    if order in st.session_state.pending_orders:
        return "待確認"
    if order in st.session_state.failed_orders:
        return "抓取失敗"
    if order in st.session_state.manual_orders:
        return "手動補卡"
    if order in st.session_state.processed_orders:
        return "已分類"
    if order in st.session_state.no_card_orders:
        return "無卡排除"
    return "未處理"


def normalize_title_for_reuse(title: str) -> str:
    """標題快取用標準化：只做保守清理，不做模糊比對。"""
    t = unicodedata.normalize("NFKC", as_text(title)).lower().strip()
    t = re.sub(r"[\s　]+", " ", t)
    t = re.sub(r"[｜|]+", "|", t)
    return t


def reusable_title_key(title: str) -> str:
    key = normalize_title_for_reuse(title)
    # 太短或太泛的標題不複用，避免錯誤擴散。
    if len(normalize_text(key)) < 10:
        return ""
    return key


def add_manual_queue(order: int, base_row: dict, reason: str = "", source_url: str = "", article_text: str = ""):
    invalidate_downloads()
    """把新聞移入手動補卡工作區；不直接輸出最終狀態。"""
    order = int(order)
    # 重要：若已存在，也要更新該列自己的標題/URL/全文，避免下方連結沿用上一筆 last_fetch。
    matched = False
    for r in st.session_state.manual_queue_rows:
        if int(r.get("原始列號", -1)) == order:
            r.update({
                "監測日期": base_row.get("監測日期", r.get("監測日期", "")),
                "訊息標題": base_row.get("訊息標題", r.get("訊息標題", "")),
                "Mastercard URL": base_row.get("Mastercard URL", r.get("Mastercard URL", "")),
                "SourceWeb URL": source_url or r.get("SourceWeb URL", "") or base_row.get("Mastercard URL", ""),
                "移入原因": reason or r.get("移入原因", ""),
                "全文": article_text or r.get("全文", ""),
            })
            matched = True
            break
    if not matched:
        st.session_state.manual_queue_rows.append({
            "原始列號": order,
            "監測日期": base_row.get("監測日期", ""),
            "訊息標題": base_row.get("訊息標題", ""),
            "Mastercard URL": base_row.get("Mastercard URL", ""),
            "SourceWeb URL": source_url or base_row.get("Mastercard URL", ""),
            "移入原因": reason,
            "全文": article_text or "",
            "內文標註": "",
        })
    sort_state_tables()
    st.session_state.manual_orders.add(order)
    st.session_state.manual_active_order = order
    st.session_state.selected_order = order
    st.session_state.no_card_orders.discard(order)
    st.session_state.failed_orders.discard(order)
    # 從無卡/失敗清單移除，避免同一筆同時在多個工作區。
    st.session_state.no_card_rows = [r for r in st.session_state.no_card_rows if int(r.get("原始列號", -1)) != order]
    st.session_state.failed_rows = [r for r in st.session_state.failed_rows if int(r.get("原始列號", -1)) != order]


def remove_manual_queue_order(order: int):
    invalidate_downloads()
    order = int(order)
    st.session_state.manual_queue_rows = [r for r in st.session_state.manual_queue_rows if int(r.get("原始列號", -1)) != order]
    st.session_state.manual_orders.discard(order)
    st.session_state.manual_staged_cards.pop(order, None)


def update_manual_queue_row(order: int, **updates):
    """更新手動補卡佇列中的單一新聞，避免標題與 SourceWeb / 全文錯位。"""
    order = int(order)
    for r in st.session_state.manual_queue_rows:
        if int(r.get("原始列號", -1)) == order:
            r.update({k: v for k, v in updates.items() if v is not None})
            return


def hydrate_manual_article(order: int, base_row: dict, queue_row: dict) -> dict:
    """確保手動補卡區顯示的是該列新聞自己的 SourceWeb 正文，而不是上一次偵測的全文。"""
    order = int(order)
    title = as_text(base_row.get("訊息標題", ""))
    mastercard_url = as_text(base_row.get("Mastercard URL", ""))
    existing_text = as_text(queue_row.get("全文", ""))
    existing_source = as_text(queue_row.get("SourceWeb URL", ""))
    if is_paper_source_label(existing_source):
        note = as_text(queue_row.get("內文標註", "")) or f"【系統標註】{PAPER_REVIEW_REASON}。"
        if not existing_text:
            existing_text = note
        update_manual_queue_row(order, **{"SourceWeb URL": PAPER_SOURCE_LABEL, "全文": existing_text, "內文標註": note})
        return {"source_url": PAPER_SOURCE_LABEL, "article_text": existing_text, "note": note}
    # 若已經有足夠全文，直接使用；避免每次重跑。
    if len(existing_text) >= 120 and existing_source:
        return {"source_url": existing_source, "article_text": existing_text, "note": as_text(queue_row.get("內文標註", ""))}
    # 優先用 Mastercard URL 重新解析 SourceWeb，確保上方標題與下方連結一致。
    input_url = mastercard_url or existing_source
    if not input_url:
        note = "【系統標註】此列沒有可用網址，請人工貼上新聞全文。"
        update_manual_queue_row(order, **{"SourceWeb URL": existing_source, "全文": existing_text, "內文標註": note})
        return {"source_url": existing_source, "article_text": existing_text, "note": note}
    try:
        source_url, article_text, error = resolve_source_and_text(input_url, title)
    except Exception as exc:
        source_url, article_text, error = existing_source or input_url, existing_text, str(exc)
    article_text = as_text(article_text)
    note = ""
    if is_paper_source_missing_error(error) or is_paper_source_label(source_url):
        note = f"【系統標註】{PAPER_REVIEW_REASON}。"
    elif error and len(article_text) < 80:
        note = f"【系統標註】SourceWeb 正文抓取不足：{error}。請開啟原始連結人工檢查。"
    if len(article_text) < 80:
        lower_blob = normalize_text(" ".join([title, article_text, as_text(source_url), as_text(input_url)]))
        if any(x in lower_blob for x in ["報紙", "剪報", "圖片", "圖像", "影像", "版面"]):
            note = "【系統標註】此筆可能為報紙圖片 / 掃描版，SourceWeb 無可解析新聞文字；請以原始連結人工檢查。"
    update_manual_queue_row(order, **{"SourceWeb URL": source_url, "全文": article_text, "內文標註": note})
    return {"source_url": source_url, "article_text": article_text, "note": note}


def add_staged_manual_card(order: int, bank: str, card: str, org: str, basis: str = "人工補卡"):
    """手動補卡先暫存；同一新聞可補多張卡，最後一次完成加入已分類。"""
    order = int(order)
    staged = st.session_state.manual_staged_cards.setdefault(order, [])
    key = normalize_key(bank, card, org)
    if not any(normalize_key(x.get("銀行別", ""), x.get("提及信用卡", ""), x.get("卡組織", "")) == key for x in staged):
        staged.append({"銀行別": bank, "提及信用卡": card, "卡組織": org, "處理狀態": "已分類", "判定依據": basis or "人工補卡"})
        return True
    return False


def stage_detected_cards_from_pasted_text(order: int, detected_df: pd.DataFrame, pending_df: pd.DataFrame | None = None) -> tuple[int, int]:
    """把貼上全文重新偵測到的卡片先放進右側暫存卡片區，不直接完成分類。

    目的：讓使用者先利用新版關鍵字表自動抓到一批候選卡，再人工補足漏掉的卡，
    最後再按「完成補卡」。避免貼上全文後仍需全部逐張手動添加。
    """
    added = 0
    skipped = 0
    frames = []
    if detected_df is not None and len(detected_df):
        tmp = detected_df.copy()
        tmp["_來源"] = "貼上全文自動偵測"
        frames.append(tmp)
    if pending_df is not None and len(pending_df):
        tmp = pending_df.copy()
        tmp["_來源"] = "貼上全文待確認候選"
        frames.append(tmp)
    if not frames:
        return 0, 0
    merged = pd.concat(frames, ignore_index=True)
    for _, r in merged.iterrows():
        bank = as_text(r.get("銀行別", ""))
        card = as_text(r.get("提及信用卡", ""))
        org = as_text(r.get("卡組織", ""))
        if not bank or not card:
            skipped += 1
            continue
        basis = as_text(r.get("判定依據", ""))
        source_label = as_text(r.get("_來源", "貼上全文自動偵測"))
        basis = f"{source_label}：{basis}" if basis else source_label
        if add_staged_manual_card(int(order), bank, card, org, basis=basis):
            added += 1
        else:
            skipped += 1
    return added, skipped


def remove_staged_manual_cards(order: int, selected_keys: list[str]):
    order = int(order)
    selected = set(selected_keys or [])
    staged = st.session_state.manual_staged_cards.get(order, [])
    st.session_state.manual_staged_cards[order] = [x for x in staged if manual_card_key(x) not in selected]


def move_manual_item_to_no_card(order: int, base_row: dict, source_url: str = "", article_text: str = "", reason: str = "人工補卡檢查後確認無卡"):
    """手動補卡檢查後確認沒有卡片時，移回無卡排除並清除暫存卡。"""
    order = int(order)
    st.session_state.last_fetch = {
        "source_url": source_url or as_text(base_row.get("Mastercard URL", "")),
        "article_text": as_text(article_text),
        "error": None,
        "字數": len(as_text(article_text)),
    }
    add_no_card(order, base_row, reason)
    remove_manual_queue_order(order)
    cache_completed_title_result(order, base_row)
    sort_state_tables()


def manual_card_key(row: dict) -> str:
    return "|".join([as_text(row.get("銀行別", "")), as_text(row.get("提及信用卡", "")), as_text(row.get("卡組織", ""))])


def classified_record_key(row: dict | pd.Series) -> str:
    """已分類結果修正用 key：同一原始列號內，同卡片視為同一筆。"""
    return "|".join([
        as_text(row.get("原始列號", "")),
        normalize_key(row.get("銀行別", ""), row.get("提及信用卡", ""), row.get("卡組織", "")),
    ])


def clear_title_cache_for_order(order: int, base_row: dict | None = None):
    """移除/修正已分類結果後，清掉同標題快取，避免錯誤卡片繼續被套用。"""
    order = int(order)
    cache = st.session_state.get("title_result_cache", {})
    delete_keys = []
    for k, v in list(cache.items()):
        try:
            if int(v.get("source_order", -999999)) == order:
                delete_keys.append(k)
        except Exception:
            continue
    if base_row:
        k = reusable_title_key(base_row.get("訊息標題", ""))
        if k:
            delete_keys.append(k)
    for k in set(delete_keys):
        cache.pop(k, None)


def get_base_row_for_order(order: int) -> dict:
    order = int(order)
    if isinstance(st.session_state.get("news_df"), pd.DataFrame) and len(st.session_state.news_df):
        rows = st.session_state.news_df[st.session_state.news_df["原始列號"].astype(str) == str(order)]
        if len(rows):
            return rows.iloc[0].to_dict()
    for list_key in ["manual_queue_rows", "pending_rows", "failed_rows", "no_card_rows"]:
        for r in st.session_state.get(list_key, []):
            try:
                if int(r.get("原始列號", -1)) == order:
                    return dict(r)
            except Exception:
                continue
    if isinstance(st.session_state.get("classified"), pd.DataFrame) and len(st.session_state.classified):
        rows = st.session_state.classified[st.session_state.classified["原始列號"] == order]
        if len(rows):
            return rows.iloc[0].to_dict()
    return {}


def refresh_completed_title_cache_for_order(order: int):
    """已分類結果被人工修正後，重新整理該列的同標題快取。"""
    order = int(order)
    base = get_base_row_for_order(order)
    clear_title_cache_for_order(order, base)
    if not base:
        return
    # 若該列仍在待確認/失敗/手動補卡，不應快取為完成結果。
    if order in st.session_state.pending_orders or order in st.session_state.failed_orders or order in st.session_state.manual_orders:
        return
    classified_left = False
    if isinstance(st.session_state.get("classified"), pd.DataFrame) and len(st.session_state.classified):
        classified_left = len(st.session_state.classified[st.session_state.classified["原始列號"] == order]) > 0
    no_card_left = any(int(r.get("原始列號", -1)) == order for r in st.session_state.get("no_card_rows", []))
    if classified_left or no_card_left:
        cache_completed_title_result(order, base)


def remove_classified_rows_by_keys(selected_keys: list[str]) -> int:
    """從已分類結果中移除選取卡片。用於修正系統誤抓或人工補卡取代舊卡。"""
    selected = set(selected_keys or [])
    if not selected or not isinstance(st.session_state.get("classified"), pd.DataFrame) or not len(st.session_state.classified):
        return 0
    invalidate_downloads()
    df = st.session_state.classified.copy()
    keys = df.apply(classified_record_key, axis=1)
    remove_mask = keys.isin(selected)
    affected_orders = set()
    if remove_mask.any():
        affected_orders = set(df.loc[remove_mask, "原始列號"].astype(int).tolist())
    removed = int(remove_mask.sum())
    st.session_state.classified = df.loc[~remove_mask].copy().reset_index(drop=True)

    # 重新整理 processed_orders，避免已經沒有卡片的列仍被視為已分類。
    remaining_orders = set()
    if len(st.session_state.classified):
        remaining_orders = set(st.session_state.classified["原始列號"].astype(int).tolist())
    st.session_state.processed_orders = set(remaining_orders)

    for order in affected_orders:
        refresh_completed_title_cache_for_order(order)
    sort_state_tables()
    return removed


def build_highlight_terms(keyword_df: pd.DataFrame, card_master: pd.DataFrame, staged_cards: list[dict] | None = None) -> list[str]:
    terms = set()
    for df, cols in [(keyword_df, ["主關鍵字", "提及信用卡"]), (card_master, ["提及信用卡"])]:
        if df is None or not len(df):
            continue
        for col in cols:
            if col in df.columns:
                for v in df[col].dropna().astype(str).tolist():
                    v = as_text(v).strip()
                    if len(normalize_text(v)) >= 2:
                        terms.add(v)
    for r in staged_cards or []:
        for col in ["提及信用卡"]:
            v = as_text(r.get(col, "")).strip()
            if len(normalize_text(v)) >= 2:
                terms.add(v)
    # 避免太多詞造成 HTML 過大；較長詞優先。
    return sorted(terms, key=lambda x: len(x), reverse=True)[:1200]


def highlighted_article_html(text: str, keyword_df: pd.DataFrame, card_master: pd.DataFrame, staged_cards: list[dict] | None = None, max_chars: int = 16000) -> str:
    raw = as_text(text)[:max_chars]
    escaped = html_lib.escape(raw)
    for term in build_highlight_terms(keyword_df, card_master, staged_cards):
        eterm = html_lib.escape(term)
        if not eterm:
            continue
        try:
            escaped = re.sub(re.escape(eterm), lambda m: f"<mark>{m.group(0)}</mark>", escaped, flags=re.IGNORECASE)
        except Exception:
            continue
    return escaped.replace("\n", "<br>")


def cache_completed_title_result(order: int, base_row: dict):
    """批量偵測加速：快取已完成結果。這不是去重，後續相同標題仍會各自輸出。"""
    key = reusable_title_key(base_row.get("訊息標題", ""))
    if not key:
        return
    order = int(order)
    # 只快取完成狀態；待確認、失敗、手動補卡不快取。
    if order in st.session_state.pending_orders or order in st.session_state.failed_orders or order in st.session_state.manual_orders:
        return
    classified = st.session_state.classified[st.session_state.classified["原始列號"] == order].to_dict("records")
    no_cards = [r for r in st.session_state.no_card_rows if int(r.get("原始列號", -1)) == order]
    if classified:
        st.session_state.title_result_cache[key] = {"source_order": order, "type": "已分類", "classified": classified, "no_card": None}
        # 若先前同標題的報紙截圖因無 SourceWeb 被放入待確認，
        # 一旦網路版完成分類，就可自動套用其卡片結果，不需逐筆人工補卡。
        apply_cached_classification_to_matching_paper_pending_orders(order, base_row)
    elif no_cards:
        st.session_state.title_result_cache[key] = {"source_order": order, "type": "無卡排除", "classified": [], "no_card": no_cards[0]}




def source_url_for_order(order: int, fallback: str = "") -> str:
    """取得指定原始列號目前工作臺中的 SourceWeb URL，避免相同標題手動補卡時沿用錯誤連結。"""
    order = int(order)
    for list_key in ["manual_queue_rows", "pending_rows", "failed_rows", "no_card_rows"]:
        for r in st.session_state.get(list_key, []):
            try:
                if int(r.get("原始列號", -1)) == order:
                    url = as_text(r.get("SourceWeb URL", ""))
                    if url:
                        return url
            except Exception:
                continue
    if isinstance(st.session_state.get("classified"), pd.DataFrame) and len(st.session_state.classified):
        rows = st.session_state.classified[st.session_state.classified["原始列號"] == order]
        if len(rows):
            url = as_text(rows.iloc[0].get("SourceWeb URL", ""))
            if url:
                return url
    return fallback


def apply_cached_classification_to_matching_paper_pending_orders(source_order: int, source_base_row: dict) -> list[int]:
    """當同標題的有 SourceWeb 網路新聞完成分類後，套用到先前待確認的報紙截圖列。

    使用情境：
    - 報紙截圖列因監測報告沒有 SourceWeb，原本必須待確認。
    - 若同標題的網路版新聞已有 SourceWeb 並已完成分類，則報紙截圖可直接套用同一組卡片。
    - 仍保留每一列自己的原始列號與監測日期，不做新聞去重。
    """
    key = reusable_title_key(source_base_row.get("訊息標題", ""))
    if not key:
        return []
    cached = st.session_state.title_result_cache.get(key)
    if not cached or cached.get("type") != "已分類" or not cached_classified_has_valid_sourceweb(cached):
        return []
    if not isinstance(st.session_state.get("news_df"), pd.DataFrame) or not len(st.session_state.news_df):
        return []

    source_order = int(source_order)
    paper_orders = []
    for r in st.session_state.get("pending_rows", []):
        try:
            order = int(r.get("原始列號", -1))
        except Exception:
            continue
        if order == source_order or order in paper_orders:
            continue
        if not is_paper_source_label(as_text(r.get("SourceWeb URL", ""))):
            continue
        if reusable_title_key(r.get("訊息標題", "")) != key:
            continue
        paper_orders.append(order)

    applied_orders: list[int] = []
    for order in paper_orders:
        raw_match = st.session_state.news_df[st.session_state.news_df["原始列號"].astype(str) == str(order)]
        if len(raw_match):
            base = raw_match.iloc[0].to_dict()
        else:
            base = next((r for r in st.session_state.pending_rows if int(r.get("原始列號", -1)) == order), {})
        if not base:
            continue

        det = pd.DataFrame(cached.get("classified", []))
        if det.empty:
            continue
        if "判定依據" in det.columns:
            det["判定依據"] = det["判定依據"].apply(
                lambda x: f"報紙截圖同標題套用第 {source_order} 列 SourceWeb 分類結果；{as_text(x)}"
            )
        else:
            det["判定依據"] = f"報紙截圖同標題套用第 {source_order} 列 SourceWeb 分類結果"

        paper_source_note = f"{PAPER_SOURCE_LABEL}；套用第 {source_order} 列同標題 SourceWeb 分類"
        st.session_state.last_fetch = {
            "source_url": paper_source_note,
            "article_text": f"【系統標註】報紙截圖 / 監測報告無 SourceWeb，已套用第 {source_order} 列同標題網路新聞分類結果。",
            "error": None,
            "字數": 0,
        }
        add_classified_rows(order, base, det)
        applied_orders.append(order)
        st.session_state.reuse_logs.append({
            "原始列號": order,
            "套用來源列號": source_order,
            "訊息標題": base.get("訊息標題", ""),
            "處理方式": "報紙截圖套用同標題 SourceWeb 分類",
        })

    if applied_orders:
        sort_state_tables()
    return applied_orders


def apply_manual_result_to_same_title_orders(source_order: int, source_base_row: dict, manual_det: pd.DataFrame, source_url: str = "", article_text: str = "") -> list[int]:
    """完成手動補卡後，自動把同標題新聞套用同一組卡片結果。

    這不是新聞去重：每一列仍以自己的原始列號輸出與計入聲量。
    用途是避免轉載新聞同標題、同卡片時，使用者必須逐篇重複手動補卡。
    """
    if manual_det is None or len(manual_det) == 0:
        return []
    key = reusable_title_key(source_base_row.get("訊息標題", ""))
    if not key or not isinstance(st.session_state.get("news_df"), pd.DataFrame) or not len(st.session_state.news_df):
        return []

    applied_orders: list[int] = []
    source_order = int(source_order)
    for _, raw_row in st.session_state.news_df.iterrows():
        try:
            order = int(raw_row.get("原始列號"))
        except Exception:
            continue
        if order == source_order:
            continue
        if reusable_title_key(raw_row.get("訊息標題", "")) != key:
            continue

        base = raw_row.to_dict()
        order_source = source_url_for_order(order, as_text(base.get("Mastercard URL", ""))) or source_url or as_text(base.get("Mastercard URL", ""))
        st.session_state.last_fetch = {
            "source_url": order_source,
            "article_text": article_text,
            "error": None,
            "字數": len(as_text(article_text)),
        }
        add_classified_rows(order, base, manual_det)
        cache_completed_title_result(order, base)
        applied_orders.append(order)
        st.session_state.reuse_logs.append({
            "原始列號": order,
            "套用來源列號": source_order,
            "訊息標題": base.get("訊息標題", ""),
            "處理方式": "手動補卡相同標題套用",
        })

    if applied_orders:
        sort_state_tables()
    return applied_orders


def preflight_paper_pending_if_needed(order: int, base_row: dict) -> bool:
    """相同標題複用前，先檢查目前列是否為無 SourceWeb 的報紙監測紀錄。

    這類列不能套用其他同標題新聞的已分類 / 無卡結果，必須獨立進待確認。
    """
    input_url = as_text(base_row.get("Mastercard URL", ""))
    title = as_text(base_row.get("訊息標題", ""))
    if not input_url:
        note = f"【系統標註】{PAPER_REVIEW_REASON}；此列未提供可解析網址。"
        paper_pending = force_paper_rows_to_pending(pd.DataFrame(), pd.DataFrame())
        st.session_state.last_fetch = {"source_url": PAPER_SOURCE_LABEL, "article_text": note, "error": PAPER_REVIEW_REASON, "字數": len(note)}
        st.session_state.last_detected = pd.DataFrame()
        st.session_state.last_pending = paper_pending
        add_pending(order, base_row, paper_pending, PAPER_SOURCE_LABEL)
        st.session_state.reuse_logs.append({"原始列號": int(order), "套用來源列號": "", "訊息標題": title, "處理方式": "報紙無 SourceWeb，強制待確認"})
        return True
    if not likely_mastercard_report_url(input_url):
        return False
    html, error = fetch_html(input_url)
    if error:
        return False
    extracted = extract_source_url_from_report_html(html)
    if extracted:
        return False
    report_text = extract_text_from_html(html, input_url)
    note = f"【系統標註】{PAPER_REVIEW_REASON}。此筆不套用相同標題既有結果；請人工開啟 Mastercard 監測報告檢查報紙圖片或掃描內容。"
    article_text = f"{note}\n\n--- Mastercard 監測報告文字 ---\n{report_text}".strip()
    detected, pending = detect_cards(title, article_text, PAPER_SOURCE_LABEL, st.session_state.keyword_df, st.session_state.card_master, st.session_state.org_rules, st.session_state.generic_terms)
    paper_pending = force_paper_rows_to_pending(detected, pending)
    st.session_state.last_fetch = {"source_url": PAPER_SOURCE_LABEL, "article_text": article_text, "error": PAPER_REVIEW_REASON, "字數": len(as_text(article_text))}
    st.session_state.last_detected = pd.DataFrame()
    st.session_state.last_pending = paper_pending
    add_pending(order, base_row, paper_pending, PAPER_SOURCE_LABEL)
    st.session_state.reuse_logs.append({"原始列號": int(order), "套用來源列號": "", "訊息標題": title, "處理方式": "報紙無 SourceWeb，強制待確認"})
    return True
def reuse_title_result_if_available(order: int, base_row: dict) -> bool:
    """若同標題已有完成結果，套用到目前列。

    v16.9 調整：
    - 監測報告無 SourceWeb 的報紙截圖，若已有同標題且有 SourceWeb 的已分類網路新聞，允許套用卡片分類。
    - 若沒有可套用的 SourceWeb 分類結果，報紙截圖仍強制進待確認並標註「報紙」。
    """
    key = reusable_title_key(base_row.get("訊息標題", ""))
    if not key:
        return False

    cached = st.session_state.title_result_cache.get(key)
    source_order = int(cached.get("source_order", -1)) if cached else -1
    if cached and source_order == int(order):
        cached = None

    # 先檢查目前列是否為「監測報告無 SourceWeb」的報紙截圖。
    # 但不同於 v16.8，現在報紙截圖可以套用同標題 SourceWeb 已分類結果。
    is_paper_missing, paper_article_text = inspect_monitoring_record_without_sourceweb(
        as_text(base_row.get("Mastercard URL", "")),
        as_text(base_row.get("訊息標題", "")),
    )

    if is_paper_missing:
        if cached and cached.get("type") == "已分類" and cached_classified_has_valid_sourceweb(cached):
            remove_order(int(order))
            det = pd.DataFrame(cached.get("classified", []))
            if det.empty:
                return False
            if "判定依據" in det.columns:
                det["判定依據"] = det["判定依據"].apply(
                    lambda x: f"報紙截圖同標題套用第 {source_order} 列 SourceWeb 分類結果；{as_text(x)}"
                )
            else:
                det["判定依據"] = f"報紙截圖同標題套用第 {source_order} 列 SourceWeb 分類結果"
            paper_source_note = f"{PAPER_SOURCE_LABEL}；套用第 {source_order} 列同標題 SourceWeb 分類"
            st.session_state.last_fetch = {
                "source_url": paper_source_note,
                "article_text": paper_article_text or f"【系統標註】報紙截圖 / 監測報告無 SourceWeb，已套用第 {source_order} 列同標題網路新聞分類結果。",
                "error": None,
                "字數": len(as_text(paper_article_text)),
            }
            add_classified_rows(int(order), base_row, det)
            st.session_state.reuse_logs.append({
                "原始列號": int(order),
                "套用來源列號": source_order,
                "訊息標題": base_row.get("訊息標題", ""),
                "處理方式": "報紙截圖套用同標題 SourceWeb 分類",
            })
            return True

        # 沒有同標題 SourceWeb 已分類結果時，仍維持報紙截圖需人工審核。
        preflight_paper_pending_if_needed(order, base_row)
        return True

    if not cached:
        return False

    remove_order(int(order))
    source_note = f"套用第 {source_order} 列相同標題結果"
    if cached.get("type") == "已分類":
        rows = []
        for r in cached.get("classified", []):
            rows.append({
                "原始列號": int(order),
                "監測日期": base_row.get("監測日期", ""),
                "訊息標題": base_row.get("訊息標題", ""),
                "Mastercard URL": base_row.get("Mastercard URL", ""),
                "SourceWeb URL": r.get("SourceWeb URL", base_row.get("Mastercard URL", "")),
                "銀行別": r.get("銀行別", ""),
                "提及信用卡": r.get("提及信用卡", ""),
                "卡組織": r.get("卡組織", ""),
                "處理狀態": "已分類",
                "判定依據": f"{source_note}；{as_text(r.get('判定依據', ''))}",
            })
        if rows:
            st.session_state.classified = pd.concat([st.session_state.classified, pd.DataFrame(rows)], ignore_index=True)
            sort_state_tables()
            st.session_state.processed_orders.add(int(order))
    elif cached.get("type") == "無卡排除":
        st.session_state.no_card_rows.append({
            "原始列號": int(order),
            "監測日期": base_row.get("監測日期", ""),
            "訊息標題": base_row.get("訊息標題", ""),
            "Mastercard URL": base_row.get("Mastercard URL", ""),
            "SourceWeb URL": cached.get("no_card", {}).get("SourceWeb URL", base_row.get("Mastercard URL", "")),
            "處理狀態": "無卡排除",
            "判定原因": source_note,
        })
        sort_state_tables()
        st.session_state.no_card_orders.add(int(order))
    else:
        return False
    st.session_state.reuse_logs.append({"原始列號": int(order), "套用來源列號": source_order, "訊息標題": base_row.get("訊息標題", "")})
    return True




def force_paper_rows_to_pending(detected_df: pd.DataFrame, pending_df: pd.DataFrame) -> pd.DataFrame:
    """報紙截圖列不自動分類；若有偵測候選，全部轉成待確認候選。"""
    frames = []
    if detected_df is not None and len(detected_df):
        tmp = detected_df.copy()
        tmp["處理狀態"] = "待確認"
        tmp["待確認原因"] = PAPER_REVIEW_REASON
        tmp["判定依據"] = tmp.get("判定依據", "").apply(lambda x: f"報紙人工審核候選；{as_text(x)}") if "判定依據" in tmp.columns else "報紙人工審核候選"
        frames.append(tmp)
    if pending_df is not None and len(pending_df):
        tmp = pending_df.copy()
        tmp["處理狀態"] = "待確認"
        tmp["待確認原因"] = tmp.get("待確認原因", "").apply(lambda x: f"{PAPER_REVIEW_REASON}；{as_text(x)}") if "待確認原因" in tmp.columns else PAPER_REVIEW_REASON
        frames.append(tmp)
    if frames:
        out = pd.concat(frames, ignore_index=True)
        return dedupe_rule_results(out)
    return pd.DataFrame([{
        "銀行別": "",
        "提及信用卡": "",
        "卡組織": "",
        "處理狀態": "待確認",
        "判定依據": PAPER_REVIEW_REASON,
        "待確認原因": PAPER_REVIEW_REASON,
    }])
def run_detection_for_row(row: pd.Series, keyword_df: pd.DataFrame, card_master: pd.DataFrame, org_rules: pd.DataFrame, generic_terms: set[str]) -> tuple[str, pd.DataFrame, pd.DataFrame, str, str | None]:
    title = as_text(row.get("訊息標題", ""))
    input_url = as_text(row.get("Mastercard URL", ""))
    source_url, article_text, error = resolve_source_and_text(input_url, title)
    # Even if error exists, try to detect from title and any fallback text.
    detected, pending = detect_cards(title, article_text, source_url, keyword_df, card_master, org_rules, generic_terms)
    return source_url, detected, pending, article_text, error


def route_detection_result(order: int, base_row: dict, source_url: str, detected_df: pd.DataFrame, pending_df: pd.DataFrame, article_text: str, error: str | None) -> str:
    """把單筆偵測結果自動分流到已分類 / 待確認 / 無卡排除 / 抓取失敗。"""
    remove_order(order)
    st.session_state.last_fetch = {"source_url": source_url, "article_text": article_text, "error": error, "字數": len(as_text(article_text))}
    st.session_state.last_detected = detected_df if detected_df is not None else pd.DataFrame()
    st.session_state.last_pending = pending_df if pending_df is not None else pd.DataFrame()

    # 報紙截圖 / 監測報告無 SourceWeb：強制進待確認，不進已分類、不進無卡排除、不進抓取失敗。
    if is_paper_source_missing_error(error) or is_paper_source_label(source_url):
        paper_pending = force_paper_rows_to_pending(st.session_state.last_detected, st.session_state.last_pending)
        st.session_state.last_detected = pd.DataFrame()
        st.session_state.last_pending = paper_pending
        add_pending(order, base_row, paper_pending, PAPER_SOURCE_LABEL)
        return f"待確認：{PAPER_REVIEW_REASON}"

    if error and len(st.session_state.last_detected) == 0 and len(st.session_state.last_pending) == 0:
        add_failed(order, base_row, source_url, error)
        return f"抓取失敗：{error}"

    if len(st.session_state.last_detected):
        add_classified_rows(order, base_row, st.session_state.last_detected)
    if len(st.session_state.last_pending):
        add_pending(order, base_row, st.session_state.last_pending, source_url)

    if len(st.session_state.last_detected) == 0 and len(st.session_state.last_pending) == 0:
        safe, reason = should_no_card(base_row.get("訊息標題", ""), article_text)
        if safe:
            add_no_card(order, base_row, reason)
            return f"無卡排除：{reason}"
        pending_row = pd.DataFrame([{
            "銀行別": "", "提及信用卡": "", "卡組織": "", "處理狀態": "待確認",
            "判定依據": reason, "待確認原因": reason,
        }])
        add_pending(order, base_row, pending_row, source_url)
        st.session_state.last_pending = pending_row
        return f"待確認：{reason}"

    parts = []
    if len(st.session_state.last_detected):
        parts.append(f"已分類 {len(st.session_state.last_detected)} 張卡")
    if len(st.session_state.last_pending):
        parts.append(f"待確認 {len(st.session_state.last_pending)} 筆")
    return "；".join(parts)


# -----------------------------------------------------------------------------
# Workbook output
# -----------------------------------------------------------------------------


def copy_row_style(ws, source_row: int, target_row: int, max_col: int):
    for c in range(1, max_col + 1):
        src = ws.cell(source_row, c)
        dst = ws.cell(target_row, c)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.font:
            dst.font = copy(src.font)
        if src.border:
            dst.border = copy(src.border)


def clear_month_sheets(wb):
    for sheet in MONTH_NAMES:
        if sheet not in wb.sheetnames:
            ws = wb.create_sheet(sheet)
            for c, h in enumerate(MONTH_COLUMNS, start=1):
                ws.cell(1, c, h)
        ws = wb[sheet]
        # Ensure header
        for c, h in enumerate(MONTH_COLUMNS, start=1):
            ws.cell(1, c, h)
        max_row = max(ws.max_row, 2)
        for r in range(2, max_row + 1):
            for c in range(1, len(MONTH_COLUMNS) + 1):
                ws.cell(r, c).value = None


def write_month_rows(wb, rows: list[dict]):
    grouped = {m: [] for m in MONTH_NAMES}
    for r in rows:
        grouped[get_month_sheet(r.get("監測日期", ""))].append(r)
    for sheet, items in grouped.items():
        ws = wb[sheet]
        for i, item in enumerate(items, start=2):
            if i > 2:
                copy_row_style(ws, 2, i, len(MONTH_COLUMNS))
            values = [
                item.get("原始列號", ""), item.get("監測日期", ""), item.get("訊息標題", ""),
                item.get("Mastercard URL", ""), item.get("SourceWeb URL", ""), item.get("處理狀態", ""),
                item.get("銀行別", ""), item.get("提及信用卡", ""), item.get("卡組織", ""),
            ]
            for c, value in enumerate(values, start=1):
                ws.cell(i, c).value = value


def refresh_summary_formulas(wb):
    if "信用卡清單_總聲量" not in wb.sheetnames:
        return
    ws = wb["信用卡清單_總聲量"]
    # Expected headers A:P
    for row in range(2, ws.max_row + 1):
        bank = as_text(ws.cell(row, 1).value)
        card = as_text(ws.cell(row, 2).value)
        if not bank or not card:
            continue
        for idx, month in enumerate(MONTH_NAMES, start=4):
            col = get_column_letter(idx)
            ws[f"{col}{row}"] = f'=COUNTIFS({month}!$F:$F,"已分類",{month}!$G:$G,$A{row},{month}!$H:$H,$B{row},{month}!$I:$I,$C{row})'
        ws[f"P{row}"] = f"=SUM(D{row}:O{row})"


def build_unprocessed_rows(news_df: pd.DataFrame) -> list[dict]:
    handled = st.session_state.processed_orders | st.session_state.no_card_orders | st.session_state.pending_orders | st.session_state.failed_orders | st.session_state.manual_orders
    rows = []
    for _, row in news_df.iterrows():
        order = int(row["原始列號"])
        if order in handled:
            continue
        rows.append({
            "原始列號": order,
            "監測日期": row.get("監測日期", ""),
            "訊息標題": row.get("訊息標題", ""),
            "Mastercard URL": row.get("Mastercard URL", ""),
            "SourceWeb URL": "",
            "處理狀態": "未處理",
            "銀行別": "", "提及信用卡": "", "卡組織": "",
        })
    return rows


def month_rows_for_output(news_df: pd.DataFrame, include_working_states: bool = False) -> list[dict]:
    """建立輸出用月份明細列。只輸出本次偵測結果，不輸出關鍵字表或設定分頁。"""
    rows = []
    for _, r in st.session_state.classified.iterrows():
        rows.append({k: r.get(k, "") for k in MONTH_COLUMNS})
    for r in st.session_state.no_card_rows:
        rows.append({k: r.get(k, "") for k in MONTH_COLUMNS})
    if include_working_states:
        for r in st.session_state.pending_rows:
            rows.append({k: r.get(k, "") for k in MONTH_COLUMNS})
        for r in st.session_state.failed_rows:
            rows.append({
                "原始列號": r.get("原始列號", ""), "監測日期": r.get("監測日期", ""), "訊息標題": r.get("訊息標題", ""),
                "Mastercard URL": r.get("Mastercard URL", ""), "SourceWeb URL": r.get("SourceWeb URL", ""),
                "處理狀態": "抓取失敗", "銀行別": "", "提及信用卡": "", "卡組織": "",
            })
        for r in st.session_state.manual_queue_rows:
            rows.append({
                "原始列號": r.get("原始列號", ""), "監測日期": r.get("監測日期", ""), "訊息標題": r.get("訊息標題", ""),
                "Mastercard URL": r.get("Mastercard URL", ""), "SourceWeb URL": r.get("SourceWeb URL", ""),
                "處理狀態": "手動補卡", "銀行別": "", "提及信用卡": "", "卡組織": "",
            })
        rows.extend(build_unprocessed_rows(news_df))
    rows = sorted(rows, key=lambda x: (get_month_sheet(x.get("監測日期", "")), int(x.get("原始列號") or 0), as_text(x.get("銀行別", "")), as_text(x.get("提及信用卡", ""))))
    return rows


def build_volume_summary_from_rows(rows: list[dict]) -> pd.DataFrame:
    """只根據輸出月份明細建立本次聲量表。"""
    classified = [r for r in rows if as_text(r.get("處理狀態")) == "已分類" and as_text(r.get("銀行別")) and as_text(r.get("提及信用卡"))]
    if not classified:
        return pd.DataFrame(columns=["銀行別", "提及信用卡", "卡組織", "聲量"])
    df = pd.DataFrame(classified)
    # 保險：同一原始列號同一卡只算一次。
    df = df.drop_duplicates(subset=["原始列號", "銀行別", "提及信用卡", "卡組織"])
    summary = df.groupby(["銀行別", "提及信用卡", "卡組織"], dropna=False).size().reset_index(name="聲量")
    return summary.sort_values(["銀行別", "聲量", "提及信用卡"], ascending=[True, False, True]).reset_index(drop=True)


def autosize_worksheet(ws, max_width: int = 46):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        values = [as_text(ws.cell(row, col).value) for row in range(1, min(ws.max_row, 300) + 1)]
        width = min(max([len(v) for v in values] + [8]) + 2, max_width)
        ws.column_dimensions[letter].width = width


def state_rows_for_export(news_df: pd.DataFrame) -> list[dict]:
    """建立可供下次上傳恢復的工作臺狀態表。

    這張表比月份明細多保留原因欄位；使用者下載暫存檔後，可再上傳此檔繼續分類。
    """
    rows: list[dict] = []
    if isinstance(st.session_state.get("classified"), pd.DataFrame) and len(st.session_state.classified):
        for _, r in st.session_state.classified.iterrows():
            rows.append({
                **{c: r.get(c, "") for c in MONTH_COLUMNS},
                "判定依據": r.get("判定依據", ""),
                "待確認原因": "", "錯誤原因": "", "判定原因": "", "移入原因": "",
            })
    for r in st.session_state.get("no_card_rows", []):
        rows.append({
            **{c: r.get(c, "") for c in MONTH_COLUMNS},
            "判定依據": "", "待確認原因": "", "錯誤原因": "",
            "判定原因": r.get("判定原因", ""), "移入原因": "",
        })
    for r in st.session_state.get("pending_rows", []):
        rows.append({
            **{c: r.get(c, "") for c in MONTH_COLUMNS},
            "判定依據": r.get("判定依據", ""),
            "待確認原因": r.get("待確認原因", ""),
            "錯誤原因": "", "判定原因": "", "移入原因": "",
        })
    for r in st.session_state.get("failed_rows", []):
        rows.append({
            "原始列號": r.get("原始列號", ""), "監測日期": r.get("監測日期", ""),
            "訊息標題": r.get("訊息標題", ""), "Mastercard URL": r.get("Mastercard URL", ""),
            "SourceWeb URL": r.get("SourceWeb URL", ""), "處理狀態": "抓取失敗",
            "銀行別": "", "提及信用卡": "", "卡組織": "",
            "判定依據": "", "待確認原因": "",
            "錯誤原因": r.get("錯誤原因", ""), "判定原因": "", "移入原因": "",
        })
    for r in st.session_state.get("manual_queue_rows", []):
        rows.append({
            "原始列號": r.get("原始列號", ""), "監測日期": r.get("監測日期", ""),
            "訊息標題": r.get("訊息標題", ""), "Mastercard URL": r.get("Mastercard URL", ""),
            "SourceWeb URL": r.get("SourceWeb URL", ""), "處理狀態": "手動補卡",
            "銀行別": "", "提及信用卡": "", "卡組織": "",
            "判定依據": "", "待確認原因": "", "錯誤原因": "", "判定原因": "",
            "移入原因": r.get("移入原因", ""),
        })
    for r in build_unprocessed_rows(news_df):
        rows.append({
            **{c: r.get(c, "") for c in MONTH_COLUMNS},
            "判定依據": "", "待確認原因": "", "錯誤原因": "", "判定原因": "", "移入原因": "",
        })
    return sorted(rows, key=lambda x: (int(x.get("原始列號") or 0), as_text(x.get("處理狀態", "")), as_text(x.get("銀行別", "")), as_text(x.get("提及信用卡", ""))))


def workbook_bytes(setting_bytes: bytes, news_df: pd.DataFrame, include_working_states: bool = False) -> bytes:
    """輸出精簡版月度結果工作簿。

    v16.1：不再把關鍵字判定表、信用卡設定清單等一併匯出。
    僅輸出：本次聲量總表 + 實際出現的月份明細表。
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    rows = month_rows_for_output(news_df, include_working_states=include_working_states)
    summary = build_volume_summary_from_rows(rows)
    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "本月聲量總表"

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(name="微軟正黑體", size=10, bold=True)
    body_font = Font(name="微軟正黑體", size=10)

    headers = ["銀行別", "提及信用卡", "卡組織", "聲量"]
    for c, h in enumerate(headers, 1):
        cell = ws_sum.cell(1, c, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    for r_idx, row in enumerate(summary.to_dict("records"), 2):
        for c_idx, h in enumerate(headers, 1):
            cell = ws_sum.cell(r_idx, c_idx, row.get(h, ""))
            cell.font = body_font
            cell.border = border
    autosize_worksheet(ws_sum)

    grouped = {}
    for r in rows:
        grouped.setdefault(get_month_sheet(r.get("監測日期", "")), []).append(r)
    if not grouped:
        grouped["月份明細"] = []
    for sheet, items in grouped.items():
        ws = wb.create_sheet(sheet if sheet in MONTH_NAMES else "月份明細")
        for c, h in enumerate(MONTH_COLUMNS, 1):
            cell = ws.cell(1, c, h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        for r_idx, item in enumerate(items, 2):
            for c_idx, h in enumerate(MONTH_COLUMNS, 1):
                cell = ws.cell(r_idx, c_idx, item.get(h, ""))
                cell.font = body_font
                cell.border = border
        autosize_worksheet(ws, max_width=60)

    if include_working_states:
        ws_state = wb.create_sheet(TEMP_STATE_SHEET)
        for c, h in enumerate(TEMP_STATE_COLUMNS, 1):
            cell = ws_state.cell(1, c, h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        for r_idx, item in enumerate(state_rows_for_export(news_df), 2):
            for c_idx, h in enumerate(TEMP_STATE_COLUMNS, 1):
                cell = ws_state.cell(r_idx, c_idx, item.get(h, ""))
                cell.font = body_font
                cell.border = border
        autosize_worksheet(ws_state, max_width=70)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


# -----------------------------------------------------------------------------
# 設定表檢查
# -----------------------------------------------------------------------------


def check_settings(card_master: pd.DataFrame, keyword_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    # cards without enabled non-generic rules
    rule_keys = set()
    for _, r in keyword_df.iterrows():
        if is_yes(r.get("是否通用詞", "N")) or as_text(r.get("類型", "")) in {"通用詞", "保護詞", "卡組織"}:
            continue
        rule_keys.add(normalize_key(r.get("銀行別", ""), r.get("提及信用卡", ""), r.get("卡組織", "")))
        # Treat blank/ALL as compatible at warning level.
        if as_text(r.get("卡組織", "")) in {"", "ALL"}:
            rule_keys.add(normalize_key(r.get("銀行別", ""), r.get("提及信用卡", ""), ""))
            rule_keys.add(normalize_key(r.get("銀行別", ""), r.get("提及信用卡", ""), "ALL"))
    missing = []
    for _, r in card_master.iterrows():
        key = normalize_key(r["銀行別"], r["提及信用卡"], r["卡組織"])
        if key not in rule_keys:
            missing.append({"銀行別": r["銀行別"], "提及信用卡": r["提及信用卡"], "卡組織": r["卡組織"], "問題": "信用卡清單有此卡，但關鍵字判定表沒有啟用規則"})

    card_keys = set(normalize_key(r["銀行別"], r["提及信用卡"], r["卡組織"]) for _, r in card_master.iterrows())
    orphan = []
    for _, r in keyword_df.iterrows():
        typ = as_text(r.get("類型", ""))
        if is_yes(r.get("是否通用詞", "N")) or typ in {"通用詞", "保護詞", "卡組織"}:
            continue
        key = normalize_key(r.get("銀行別", ""), r.get("提及信用卡", ""), r.get("卡組織", ""))
        if key not in card_keys:
            orphan.append({"銀行別": r.get("銀行別", ""), "主關鍵字": r.get("主關鍵字", ""), "提及信用卡": r.get("提及信用卡", ""), "卡組織": r.get("卡組織", ""), "問題": "關鍵字規則對應不到信用卡清單_總聲量"})
    return pd.DataFrame(missing), pd.DataFrame(orphan)


# -----------------------------------------------------------------------------
# UI
# -----------------------------------------------------------------------------


st.set_page_config(page_title="信用卡聲量分析系統", layout="wide", initial_sidebar_state="collapsed")
init_state()

st.markdown(
    """
    <style>
    .main-title {font-size: 30px; font-weight: 800; margin-bottom: 4px;}
    .subtle {color:#666; font-size: 14px;}
    .step-guide {color:#777; font-size: 13px; line-height: 1.55; margin-top: 2px; margin-bottom: 12px;}
    .step-title {font-size: 20px; font-weight: 800; margin: 20px 0 6px 0;}
    .step-caption {color:#777; font-size: 13px; margin: -2px 0 10px 0;}
    .status-strip {border:1px solid #ddd; border-radius:10px; padding:10px 12px; background:#fafafa; margin:8px 0 14px 0;}
    .result-chip {display:inline-block; padding:5px 9px; border-radius:16px; margin:3px 4px 3px 0; border:1px solid #ddd; background:#f7f7f7;}
    .danger {color:#b00020; font-weight:700;}
    .article-scroll-box {
        max-height: 520px;
        overflow-y: auto;
        overflow-x: hidden;
        white-space: pre-wrap;
        word-break: break-word;
        border: 1px solid #ddd;
        border-radius: 10px;
        padding: 12px 14px;
        background: #fff;
        line-height: 1.65;
        font-size: 14px;
    }
    .link-row {margin: 2px 0 6px 0; font-size: 14px; word-break: break-all;}
    .link-row a {word-break: break-all;}
    mark {background: #fff3a3; padding: 0 2px; border-radius: 2px;}
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown(
    f"""
    <div class='main-title'>信用卡聲量分析系統</div>
    <div class='subtle'>{APP_VERSION}</div>
    <div class='step-guide'>
    Step 1 上傳檔案或恢復進度 → Step 2 執行單一 / 批量偵測 → Step 3 在下方表格處理待確認、無卡排除與手動補卡 → Step 4 下載暫存檔或完成版。
    </div>
    """,
    unsafe_allow_html=True,
)

st.markdown("<div class='step-title'>Step 1｜上傳檔案 / 恢復進度</div>", unsafe_allow_html=True)
with st.expander("上傳檔案", expanded=True):
    c1, c2, c3 = st.columns(3)
    with c1:
        raw_file = st.file_uploader("① 每月 Mastercard raw data", type=["xlsx", "xls"], key="raw_file_v13")
    with c2:
        setting_file = st.file_uploader("② 固定的信用卡聲量 + 關鍵字判定表", type=["xlsx"], key="setting_file_v13")
    with c3:
        resume_workbook_file = st.file_uploader("③ 可選：上次下載的暫存月度結果表", type=["xlsx"], key="resume_workbook_v174")
    st.caption("新月份：raw data + 設定表；續作：暫存月度結果表 + 設定表。")

saved_payload = load_saved_progress()
with st.expander("進度", expanded=False):
    st.caption(saved_progress_label(saved_payload))
    cc1, cc2, cc3 = st.columns(3)
    with cc1:
        if st.button("恢復上次進度", disabled=not bool(saved_payload), use_container_width=True, key="btn_restore_progress_v172"):
            if restore_progress_to_state(saved_payload):
                st.success("已恢復上次進度。")
                save_progress_to_disk("restore")
                st.rerun()
            else:
                st.error("恢復失敗，請重新上傳檔案。")
    with cc2:
        if st.button("立即保存目前進度", disabled=not isinstance(st.session_state.get("news_df"), pd.DataFrame) or st.session_state.get("news_df").empty, use_container_width=True, key="btn_save_progress_v172"):
            if save_progress_to_disk("manual"):
                st.success("已保存目前進度。")
            else:
                st.warning("目前沒有可保存的工作進度。")
    with cc3:
        if st.button("清除已保存進度", disabled=not bool(saved_payload), use_container_width=True, key="btn_clear_progress_v172"):
            clear_saved_progress_file()
            st.warning("已清除已保存進度。")
            st.rerun()
    if st.session_state.get("last_autosave_error"):
        st.caption(f"保存錯誤：{st.session_state.last_autosave_error}")

use_temp_workbook_resume = bool(resume_workbook_file and setting_file)
if resume_workbook_file and setting_file:
    raw_bytes = resume_workbook_file.getvalue()
    setting_bytes = setting_file.getvalue()
    raw_sig = "temp_resume|" + file_signature(getattr(resume_workbook_file, "name", "temp_resume"), raw_bytes)
    setting_sig = file_signature(getattr(setting_file, "name", "setting"), setting_bytes)
    st.info("已選擇以『暫存月度結果表』接續上次進度；系統會讀取已分類、待確認、抓取失敗、手動補卡與未處理狀態。")
elif raw_file and setting_file:
    raw_bytes = raw_file.getvalue()
    setting_bytes = setting_file.getvalue()
    raw_sig = file_signature(getattr(raw_file, "name", "raw"), raw_bytes)
    setting_sig = file_signature(getattr(setting_file, "name", "setting"), setting_bytes)
elif st.session_state.get("raw_bytes") and st.session_state.get("setting_bytes") and isinstance(st.session_state.get("news_df"), pd.DataFrame) and len(st.session_state.news_df):
    raw_bytes = st.session_state.raw_bytes
    setting_bytes = st.session_state.setting_bytes
    raw_sig = st.session_state.raw_sig
    setting_sig = st.session_state.setting_sig
    use_temp_workbook_resume = bool(st.session_state.get("loaded_from_temp_workbook", False))
    st.info("目前使用已恢復的上次進度；若要更換月份或設定表，請重新上傳檔案。")
else:
    st.info("請先上傳 raw data + 設定表，或上傳暫存月度結果表 + 設定表；也可在「進度保存 / 恢復」中恢復上次進度。")
    st.stop()

# 上傳檔案有更換時才重新讀取設定；一般按鈕操作只用 session_state 中的資料，避免每次重跑都重讀 Excel。
files_changed = (st.session_state.get("raw_sig") != raw_sig) or (st.session_state.get("setting_sig") != setting_sig)
if files_changed:
    reset_working_state("已偵測到上傳檔案變更，工作臺狀態已重置。")
    st.session_state.raw_sig = raw_sig
    st.session_state.setting_sig = setting_sig
    st.session_state.raw_bytes = raw_bytes
    st.session_state.setting_bytes = setting_bytes
    try:
        st.session_state.card_master = load_card_master(setting_bytes)
        st.session_state.keyword_df = load_keyword_rules(setting_bytes)
        st.session_state.org_rules = load_org_rules(setting_bytes)
        st.session_state.generic_terms = load_generic_terms(setting_bytes)
        st.session_state.missing_rules_df, st.session_state.orphan_rules_df = check_settings(st.session_state.card_master, st.session_state.keyword_df)
        if use_temp_workbook_resume:
            progress = read_temp_workbook_progress(raw_bytes)
            restore_temp_workbook_progress_to_state(progress)
            st.session_state.loaded_from_temp_workbook = True
            st.session_state.last_message = "已從暫存月度結果表恢復工作進度，可繼續處理待確認、抓取失敗與未處理項目。"
        else:
            st.session_state.news_df = read_raw_news_cached(raw_bytes, getattr(raw_file, "name", "raw.xlsx"))
            st.session_state.loaded_from_temp_workbook = False
    except Exception as e:
        st.error("檔案讀取失敗。請確認分頁名稱、欄位或暫存月度結果表格式是否符合定案格式。")
        st.exception(e)
        st.stop()

news_df = st.session_state.news_df
card_master = st.session_state.card_master
keyword_df = st.session_state.keyword_df
org_rules = st.session_state.org_rules
generic_terms = st.session_state.generic_terms
missing_rules_df = st.session_state.missing_rules_df
orphan_rules_df = st.session_state.orphan_rules_df
setting_bytes = st.session_state.setting_bytes

# Metrics
handled_orders = st.session_state.processed_orders | st.session_state.no_card_orders | st.session_state.pending_orders | st.session_state.failed_orders | st.session_state.manual_orders
unprocessed_count = max(0, len(news_df) - len(handled_orders))
metrics = st.columns(6)
metrics[0].metric("raw data 列數", len(news_df))
metrics[1].metric("已分類", len(st.session_state.processed_orders))
metrics[2].metric("無卡排除", len(st.session_state.no_card_orders))
metrics[3].metric("待確認", len(st.session_state.pending_orders))
metrics[4].metric("抓取失敗", len(st.session_state.failed_orders))
metrics[5].metric("未處理", unprocessed_count)

status_text = f"信用卡清單 {len(card_master)} 筆｜啟用關鍵字規則 {len(keyword_df)} 筆｜缺少關鍵字規則的卡 {len(missing_rules_df)} 筆｜關鍵字對不到聲量表 {len(orphan_rules_df)} 筆"
st.markdown(f"<div class='status-strip'><b>設定表狀態：</b>{status_text}</div>", unsafe_allow_html=True)

st.markdown("<div class='step-title'>Step 2｜執行新聞偵測</div>", unsafe_allow_html=True)
st.markdown("<div class='step-caption'>選擇新聞後可單筆偵測，或從目前列開始批量偵測。</div>", unsafe_allow_html=True)

# News selector
news_options = []
order_map = {}
for _, r in news_df.iterrows():
    order = int(r["原始列號"])
    option = f"{order}. [{status_for(order)}] {as_text(r['訊息標題'])[:80]}"
    news_options.append(option)
    order_map[option] = order

if st.session_state.selected_order is None and len(news_df):
    st.session_state.selected_order = int(news_df.iloc[0]["原始列號"])

col_left, col_right = st.columns([1.35, 1])
with col_left:
    current_idx = 0
    if st.session_state.selected_order:
        for i, opt in enumerate(news_options):
            if order_map[opt] == st.session_state.selected_order:
                current_idx = i
                break
    selected_opt = st.selectbox("選擇新聞", news_options, index=current_idx, key="news_selector_v143")
    selected_order = order_map[selected_opt]
    st.session_state.selected_order = selected_order
    manual_url = st.text_input("或貼上 Mastercard / 原始新聞網址覆蓋本筆 URL", value="", placeholder="https://...")

with col_right:
    batch_size = st.selectbox("批次處理筆數", [10, 20, 50, 100], index=1)
    st.caption("批次從目前選取新聞開始；相同標題會自動套用前面已完成結果，不會刪除新聞列。")

base_row = news_df[news_df["原始列號"] == selected_order].iloc[0].to_dict()
if manual_url.strip():
    base_row["Mastercard URL"] = manual_url.strip()

b1, b2 = st.columns(2)
with b1:
    batch_clicked = st.button("批量偵測", use_container_width=True, type="primary", key="btn_batch_detect")
with b2:
    detect_clicked = st.button("單一偵測", use_container_width=True, key="btn_single_detect")

if detect_clicked:
    source_url, detected_df, pending_df, article_text, error = run_detection_for_row(pd.Series(base_row), keyword_df, card_master, org_rules, generic_terms)
    message = route_detection_result(selected_order, base_row, source_url, detected_df, pending_df, article_text, error)
    cache_completed_title_result(selected_order, base_row)
    st.success(f"單一偵測完成：{message}")

if batch_clicked:
    start_idx = int(news_df.index[news_df["原始列號"] == selected_order][0])
    effective = 0
    scanned = 0
    reused = 0
    progress = st.progress(0, text=f"已偵測 0/{int(batch_size)}")
    progress_note = st.empty()
    for i in range(start_idx, len(news_df)):
        if effective >= int(batch_size):
            break
        row = news_df.iloc[i]
        order = int(row["原始列號"])
        if status_for(order) != "未處理":
            continue
        scanned += 1
        base = row.to_dict()
        if reuse_title_result_if_available(order, base):
            effective += 1
            reused += 1
            progress.progress(min(1.0, effective / max(1, int(batch_size))), text=f"已偵測 {effective}/{int(batch_size)}")
            progress_note.caption(f"第 {order} 列：套用相同標題既有結果。")
            continue
        source_url, detected_df, pending_df, article_text, error = run_detection_for_row(row, keyword_df, card_master, org_rules, generic_terms)
        message = route_detection_result(order, base, source_url, detected_df, pending_df, article_text, error)
        if str(message).startswith("抓取失敗"):
            progress_note.caption(f"第 {order} 列抓取失敗，不計入有效處理數：{message}")
            continue
        effective += 1
        cache_completed_title_result(order, base)
        progress.progress(min(1.0, effective / max(1, int(batch_size))), text=f"已偵測 {effective}/{int(batch_size)}")
        progress_note.caption(f"第 {order} 列：{message}")
    st.success(f"批次完成：有效處理 {effective}/{int(batch_size)}；自動套用相同標題結果 {reused} 筆；抓取失敗不計入有效筆數。")

# Results panel
st.markdown("**本次偵測結果**")
res1, res2 = st.columns(2)
with res1:
    st.markdown("**高信心 / 已分類**")
    if len(st.session_state.last_detected):
        st.dataframe(st.session_state.last_detected[["銀行別", "提及信用卡", "卡組織", "判定依據"]], use_container_width=True, height=180)
    else:
        st.caption("無")
with res2:
    st.markdown("**待確認**")
    if len(st.session_state.last_pending):
        st.dataframe(st.session_state.last_pending[["銀行別", "提及信用卡", "卡組織", "待確認原因", "判定依據"]], use_container_width=True, height=180)
    else:
        st.caption("無")

if st.session_state.reuse_logs:
    with st.expander("相同標題結果複用紀錄", expanded=False):
        st.dataframe(pd.DataFrame(st.session_state.reuse_logs[-100:]), use_container_width=True, height=220)

# Download status only; compact download panel is placed at the bottom to keep the main workspace short.
unfinished = len(st.session_state.pending_orders) + len(st.session_state.failed_orders) + len(st.session_state.manual_orders) + unprocessed_count

# Workbench tabs
st.divider()
st.markdown("<div class='step-title'>Step 3｜人工分類工作臺</div>", unsafe_allow_html=True)
st.markdown("<div class='step-caption'>主要處理區：待確認、抓取失敗、無卡排除與手動補卡。下方表格勾選後可直接移動分類，不會重新跑整批偵測。</div>", unsafe_allow_html=True)
tabs = st.tabs(["待確認", "抓取失敗", "無卡排除", "全文預覽 + 手動補卡", "設定表檢查", "完整結果"])

with tabs[0]:
    pending_df = pd.DataFrame(st.session_state.pending_rows)
    if len(pending_df):
        editable = pending_df.copy()
        editable.insert(0, "選取", False)
        show_cols = ["選取", "原始列號", "訊息標題", "銀行別", "提及信用卡", "卡組織", "待確認原因", "判定依據"]
        with st.form("pending_action_form_v15"):
            edited_pending = st.data_editor(
                editable[show_cols],
                use_container_width=True,
                height=340,
                key="pending_select_editor_v15",
                column_config={"選取": st.column_config.CheckboxColumn("選取")},
                disabled=[c for c in show_cols if c != "選取"],
            )
            selected_pending = edited_pending[edited_pending["選取"] == True].copy()
            p1, p2, p3 = st.columns(3)
            with p1:
                submit_to_classified = st.form_submit_button("加入已分類", use_container_width=True)
            with p2:
                submit_to_no_card = st.form_submit_button("加入無卡排除", use_container_width=True)
            with p3:
                submit_to_manual = st.form_submit_button("加入手動補卡", use_container_width=True)

        selected_keys = set(pending_record_key(r) for _, r in selected_pending.iterrows())
        if submit_to_classified and not selected_keys:
            st.warning("請先勾選待確認項目。")
        if submit_to_classified and selected_keys:
            full_pending = pd.DataFrame(st.session_state.pending_rows)
            selected_full = full_pending[full_pending.apply(lambda r: pending_record_key(r) in selected_keys, axis=1)]
            for order, grp in selected_full.groupby("原始列號"):
                base = news_df[news_df["原始列號"] == int(order)].iloc[0].to_dict()
                add_classified_rows(int(order), base, grp, clear_pending_order=False)
                cache_completed_title_result(int(order), base)
            st.session_state.pending_rows = [r for r in st.session_state.pending_rows if pending_record_key(r) not in selected_keys]
            refresh_pending_orders()
            save_progress_to_disk("auto"); st.rerun()

        if submit_to_no_card and not selected_keys:
            st.warning("請先勾選待確認項目。")
        if submit_to_no_card and selected_keys:
            affected_orders = sorted(set(selected_pending["原始列號"].astype(int).tolist()))
            st.session_state.pending_rows = [r for r in st.session_state.pending_rows if pending_record_key(r) not in selected_keys]
            refresh_pending_orders()
            for order in affected_orders:
                has_classified = len(st.session_state.classified[st.session_state.classified["原始列號"] == order]) > 0
                has_pending = any(int(r.get("原始列號", -1)) == order for r in st.session_state.pending_rows)
                if not has_classified and not has_pending:
                    base = news_df[news_df["原始列號"] == order].iloc[0].to_dict()
                    add_no_card(order, base, "待確認項目人工選取加入無卡排除")
                    cache_completed_title_result(order, base)
            save_progress_to_disk("auto"); st.rerun()

        if submit_to_manual and not selected_keys:
            st.warning("請先勾選待確認項目。")
        if submit_to_manual and selected_keys:
            full_pending = pd.DataFrame(st.session_state.pending_rows)
            selected_full = full_pending[full_pending.apply(lambda r: pending_record_key(r) in selected_keys, axis=1)]
            for order, grp in selected_full.groupby("原始列號"):
                order = int(order)
                base = news_df[news_df["原始列號"] == order].iloc[0].to_dict()
                source_url = as_text(grp.iloc[0].get("SourceWeb URL", ""))
                add_manual_queue(order, base, "待確認項目移入手動補卡", source_url=source_url)
            st.session_state.pending_rows = [r for r in st.session_state.pending_rows if pending_record_key(r) not in selected_keys]
            refresh_pending_orders()
            save_progress_to_disk("auto"); st.rerun()
    else:
        st.info("目前沒有待確認資料。")

with tabs[1]:
    failed_df = pd.DataFrame(st.session_state.failed_rows)
    if len(failed_df):
        failed_view = failed_df.copy()
        failed_view.insert(0, "選取", False)
        with st.form("failed_action_form_v15"):
            edited_failed = st.data_editor(
                failed_view,
                use_container_width=True,
                height=260,
                key="failed_select_editor_v15",
                column_config={"選取": st.column_config.CheckboxColumn("選取")},
                disabled=[c for c in failed_view.columns if c != "選取"],
            )
            selected_failed = edited_failed[edited_failed["選取"] == True]
            submit_failed_to_manual = st.form_submit_button("選取項目加入手動補卡", use_container_width=True)
        if submit_failed_to_manual and len(selected_failed) == 0:
            st.warning("請先勾選抓取失敗項目。")
        if submit_failed_to_manual and len(selected_failed):
            for _, r in selected_failed.iterrows():
                order = int(r["原始列號"])
                base = news_df[news_df["原始列號"] == order].iloc[0].to_dict()
                add_manual_queue(order, base, f"抓取失敗移入手動補卡：{as_text(r.get('錯誤原因',''))}", source_url=as_text(r.get("SourceWeb URL", "")))
            save_progress_to_disk("auto"); st.rerun()
    else:
        st.info("目前沒有抓取失敗資料。")

with tabs[2]:
    no_card_df = pd.DataFrame(st.session_state.no_card_rows)
    if len(no_card_df):
        editable_no = no_card_df.copy()
        editable_no.insert(0, "選取", False)
        with st.form("no_card_action_form_v15"):
            edited_no = st.data_editor(
                editable_no,
                use_container_width=True,
                height=300,
                key="no_card_select_editor_v15",
                column_config={"選取": st.column_config.CheckboxColumn("選取")},
                disabled=[c for c in editable_no.columns if c != "選取"],
            )
            selected_no = edited_no[edited_no["選取"] == True]
            submit_no_to_manual = st.form_submit_button("加入手動補卡", use_container_width=True)
        if submit_no_to_manual and len(selected_no) == 0:
            st.warning("請先勾選無卡排除項目。")
        if submit_no_to_manual and len(selected_no):
            for _, r in selected_no.iterrows():
                order = int(r["原始列號"])
                base = news_df[news_df["原始列號"] == order].iloc[0].to_dict()
                add_manual_queue(order, base, "無卡排除改為手動補卡", source_url=as_text(r.get("SourceWeb URL", "")))
            save_progress_to_disk("auto"); st.rerun()
    else:
        st.info("目前沒有無卡排除資料。")

with tabs[3]:
    st.caption("在此查看 SourceWeb 全文並暫存多張人工補卡；確認同一篇新聞的所有卡片都補完後，再一次加入已分類。")
    manual_df = pd.DataFrame(st.session_state.manual_queue_rows)
    target_order = selected_order
    if len(manual_df):
        manual_options = [f"{int(r['原始列號'])}. {as_text(r['訊息標題'])[:70]}" for _, r in manual_df.iterrows()]
        active_order = int(st.session_state.get("manual_active_order") or int(manual_df.iloc[0]["原始列號"]))
        manual_index = 0
        for idx, opt in enumerate(manual_options):
            try:
                if int(opt.split('.', 1)[0]) == active_order:
                    manual_index = idx
                    break
            except Exception:
                pass
        chosen_manual = st.selectbox("手動補卡清單", manual_options, index=manual_index, key="manual_queue_selector_v163")
        target_order = int(chosen_manual.split('.', 1)[0])
        st.session_state.manual_active_order = target_order
        st.session_state.selected_order = target_order
    else:
        st.info("手動補卡清單目前為空；下方會使用目前選取新聞。")

    target_base = news_df[news_df["原始列號"] == int(target_order)].iloc[0].to_dict()
    queue_row = next((r for r in st.session_state.manual_queue_rows if int(r.get("原始列號", -1)) == int(target_order)), {})

    hydrated = hydrate_manual_article(int(target_order), target_base, queue_row) if queue_row else {"source_url": as_text(target_base.get("Mastercard URL", "")), "article_text": "", "note": ""}
    queue_row = next((r for r in st.session_state.manual_queue_rows if int(r.get("原始列號", -1)) == int(target_order)), queue_row)
    preview_text = as_text(hydrated.get("article_text", ""))
    preview_source = as_text(hydrated.get("source_url", "")) or as_text(target_base.get("Mastercard URL", ""))
    preview_note = as_text(hydrated.get("note", "")) or as_text(queue_row.get("內文標註", ""))

    st.markdown(f"**目前處理：第 {int(target_order)} 列｜{as_text(target_base.get('訊息標題',''))}**")
    st.markdown(
        f"<div class='link-row'><b>Mastercard 監測連結：</b>{markdown_link(target_base.get('Mastercard URL', ''), target_base.get('Mastercard URL', ''))}</div>"
        f"<div class='link-row'><b>SourceWeb 原始新聞：</b>{markdown_link(preview_source, preview_source)}</div>"
        f"<div class='link-row'><b>目前全文字數：</b>{len(preview_text)}｜<b>移入原因：</b>{html_lib.escape(as_text(queue_row.get('移入原因', '')))}</div>",
        unsafe_allow_html=True,
    )
    if preview_note:
        st.warning(preview_note)

    # 已分類結果修正：同一新聞若系統原本已分類錯卡，人工補卡時可以先移除錯誤卡片。
    existing_classified_for_order = pd.DataFrame()
    if isinstance(st.session_state.get("classified"), pd.DataFrame) and len(st.session_state.classified):
        existing_classified_for_order = st.session_state.classified[
            st.session_state.classified["原始列號"] == int(target_order)
        ].copy()
    if len(existing_classified_for_order):
        st.markdown("**本列目前已分類卡片（可移除錯誤卡片）**")
        fix_df = existing_classified_for_order.copy()
        fix_df.insert(0, "選取", False)
        fix_cols = ["選取", "原始列號", "銀行別", "提及信用卡", "卡組織", "判定依據"]
        with st.form(key=f"classified_fix_form_{int(target_order)}"):
            edited_fix = st.data_editor(
                fix_df[fix_cols],
                use_container_width=True,
                hide_index=True,
                height=160,
                key=f"classified_fix_editor_{int(target_order)}",
                column_config={"選取": st.column_config.CheckboxColumn("選取")},
                disabled=[c for c in fix_cols if c != "選取"],
            )
            remove_existing_classified = st.form_submit_button("移除選取已分類卡片", use_container_width=True)
        if remove_existing_classified:
            selected_existing_keys = [
                classified_record_key(r)
                for _, r in edited_fix[edited_fix["選取"] == True].iterrows()
            ] if "選取" in edited_fix.columns else []
            if not selected_existing_keys:
                st.warning("請先勾選要移除的已分類卡片。")
            else:
                removed_count = remove_classified_rows_by_keys(selected_existing_keys)
                st.session_state.last_message = f"第 {int(target_order)} 列已移除 {removed_count} 筆已分類卡片。"
                save_progress_to_disk("auto"); st.rerun()
    else:
        st.caption("本列目前沒有已分類卡片；可直接在右側暫存人工補卡。")

    left, right = st.columns([1.25, 1])
    with left:
        pasted_text = st.text_area("SourceWeb 全文 / 可貼上人工全文", value=preview_text[:16000], height=260, key=f"manual_text_{int(target_order)}")
        staged_cards = st.session_state.manual_staged_cards.get(int(target_order), [])
        st.markdown("**螢光筆標記全文中的卡片關鍵字**")
        if pasted_text.strip():
            st.markdown(
                f"<div class='article-scroll-box'>{highlighted_article_html(pasted_text, keyword_df, card_master, staged_cards, max_chars=30000)}</div>",
                unsafe_allow_html=True,
            )
        else:
            st.info("目前沒有可顯示的新聞全文。若 SourceWeb 為報紙圖片或無法解析，請貼上人工全文。")

        col_redetect, col_refresh = st.columns(2)
        with col_redetect:
            if st.button("用貼上全文重新偵測", use_container_width=True, disabled=not pasted_text.strip(), key="btn_redetect_pasted_text_v163"):
                det, pen = detect_cards(target_base.get("訊息標題", ""), pasted_text, preview_source, keyword_df, card_master, org_rules, generic_terms)
                st.session_state.last_fetch = {"source_url": preview_source or "人工貼全文", "article_text": pasted_text, "error": None, "字數": len(pasted_text)}
                update_manual_queue_row(int(target_order), **{"SourceWeb URL": preview_source or "人工貼全文", "全文": pasted_text, "內文標註": ""})
                added_count, skipped_count = stage_detected_cards_from_pasted_text(int(target_order), det, pen)
                if added_count:
                    st.session_state.last_message = (
                        f"第 {int(target_order)} 列已用貼上全文重新偵測，"
                        f"並先加入 {added_count} 張候選卡到右側暫存卡片區。"
                        f"請確認是否還要補卡或移除候選卡，再按『完成補卡，全部加入已分類』。"
                    )
                else:
                    st.session_state.last_message = (
                        f"第 {int(target_order)} 列已用貼上全文重新偵測，但沒有新增候選卡片。"
                        f"請人工補卡，或確認無卡後移回無卡排除。"
                    )
                save_progress_to_disk("auto"); st.rerun()
        with col_refresh:
            if st.button("重新載入 SourceWeb 全文", use_container_width=True, key="btn_refresh_sourceweb_manual_v163"):
                update_manual_queue_row(int(target_order), **{"全文": "", "內文標註": ""})
                save_progress_to_disk("auto"); st.rerun()

    with right:
        st.markdown("**暫存本篇要補的卡片**")
        banks = sorted(card_master["銀行別"].dropna().unique().tolist())
        bank = st.selectbox("銀行別", banks, key="manual_bank_v163")
        cards_for_bank = card_master[card_master["銀行別"] == bank]
        card = st.selectbox("提及信用卡", cards_for_bank["提及信用卡"].dropna().unique().tolist(), key="manual_card_v163")
        orgs = cards_for_bank[cards_for_bank["提及信用卡"] == card]["卡組織"].fillna("").astype(str).tolist()
        org = st.selectbox("卡組織", orgs if orgs else [""], key="manual_org_v163")
        if st.button("加入本篇暫存補卡清單", use_container_width=True, key="btn_stage_manual_card_v163"):
            add_staged_manual_card(int(target_order), bank, card, org)
            st.session_state.last_message = f"已暫存：{bank}｜{card}｜{org}"
            save_progress_to_disk("auto"); st.rerun()

        staged_cards = st.session_state.manual_staged_cards.get(int(target_order), [])
        if staged_cards:
            staged_df = pd.DataFrame(staged_cards)
            staged_df.insert(0, "選取", False)
            with st.form(key=f"manual_staged_form_{int(target_order)}"):
                edited_staged = st.data_editor(staged_df, use_container_width=True, hide_index=True, height=200, key=f"manual_staged_editor_{int(target_order)}")
                c1, c2 = st.columns(2)
                with c1:
                    remove_selected = st.form_submit_button("移除選取暫存卡", use_container_width=True)
                with c2:
                    finish_manual = st.form_submit_button("完成補卡，全部加入已分類", use_container_width=True)
            selected_remove = []
            if "選取" in edited_staged.columns:
                selected_remove = [manual_card_key(r) for _, r in edited_staged[edited_staged["選取"] == True].iterrows()]
            if remove_selected:
                if not selected_remove:
                    st.warning("請先勾選要移除的暫存卡。")
                else:
                    remove_staged_manual_cards(int(target_order), selected_remove)
                    save_progress_to_disk("auto"); st.rerun()
            if finish_manual:
                current_staged = st.session_state.manual_staged_cards.get(int(target_order), [])
                if not current_staged:
                    st.warning("請先加入至少一張卡片。")
                else:
                    manual_det = pd.DataFrame(current_staged)
                    st.session_state.last_fetch = {"source_url": preview_source, "article_text": pasted_text, "error": None, "字數": len(as_text(pasted_text))}
                    add_classified_rows(int(target_order), target_base, manual_det)
                    cache_completed_title_result(int(target_order), target_base)
                    applied_same_title = apply_manual_result_to_same_title_orders(
                        int(target_order),
                        target_base,
                        manual_det,
                        preview_source,
                        pasted_text,
                    )
                    remove_manual_queue_order(int(target_order))
                    if applied_same_title:
                        st.session_state.last_message = (
                            f"第 {int(target_order)} 列已完成補卡，共加入 {len(current_staged)} 張卡；"
                            f"並自動套用到 {len(applied_same_title)} 筆相同標題新聞："
                            f"{', '.join(map(str, applied_same_title[:20]))}"
                            f"{'...' if len(applied_same_title) > 20 else ''}。"
                        )
                    else:
                        st.session_state.last_message = f"第 {int(target_order)} 列已完成補卡，共加入 {len(current_staged)} 張卡。"
                    save_progress_to_disk("auto"); st.rerun()
        else:
            st.info("尚未暫存卡片。若本篇有多張卡，請逐張加入暫存清單，全部補完後再按完成。")

        st.divider()
        st.markdown("**確認本篇沒有信用卡**")
        st.caption("若人工檢查後確認這篇新聞沒有任何應計入的信用卡，可直接移回無卡排除；若已暫存卡片，按下後會清空本篇暫存卡。")
        if st.button("確認無卡，移回無卡排除", use_container_width=True, key=f"btn_manual_back_to_no_card_{int(target_order)}"):
            move_manual_item_to_no_card(
                int(target_order),
                target_base,
                preview_source,
                pasted_text,
                "人工補卡檢查後確認無卡",
            )
            st.session_state.last_message = f"第 {int(target_order)} 列已移回無卡排除。"
            save_progress_to_disk("auto"); st.rerun()

with tabs[4]:
    st.markdown("**信用卡清單有、但關鍵字判定表沒有啟用規則**")
    st.dataframe(missing_rules_df, use_container_width=True, height=220)
    st.markdown("**關鍵字判定表有、但對不到信用卡清單_總聲量**")
    st.dataframe(orphan_rules_df, use_container_width=True, height=220)

with tabs[5]:
    st.markdown("**已分類卡片結果**")
    st.caption("若系統誤抓或人工補卡後同一則新聞多出錯誤卡片，可在這裡勾選並移除。")
    sort_state_tables()
    if isinstance(st.session_state.get("classified"), pd.DataFrame) and len(st.session_state.classified):
        classified_edit = st.session_state.classified.copy()
        classified_edit.insert(0, "選取", False)
        result_cols = ["選取", "原始列號", "監測日期", "訊息標題", "銀行別", "提及信用卡", "卡組織", "SourceWeb URL", "判定依據"]
        result_cols = [c for c in result_cols if c in classified_edit.columns]
        with st.form("classified_result_remove_form_v171"):
            edited_result = st.data_editor(
                classified_edit[result_cols],
                use_container_width=True,
                height=320,
                key="classified_result_remove_editor_v171",
                column_config={"選取": st.column_config.CheckboxColumn("選取")},
                disabled=[c for c in result_cols if c != "選取"],
            )
            remove_result_rows = st.form_submit_button("移除選取已分類卡片", use_container_width=True)
        if remove_result_rows:
            selected_result_keys = [
                classified_record_key(r)
                for _, r in edited_result[edited_result["選取"] == True].iterrows()
            ] if "選取" in edited_result.columns else []
            if not selected_result_keys:
                st.warning("請先勾選要移除的已分類卡片。")
            else:
                removed_count = remove_classified_rows_by_keys(selected_result_keys)
                st.session_state.last_message = f"已移除 {removed_count} 筆已分類卡片。"
                save_progress_to_disk("auto"); st.rerun()
    else:
        st.info("目前沒有已分類卡片結果。")
    with st.expander("raw data 預覽", expanded=False):
        preview = news_df.copy()
        preview["處理狀態"] = preview["原始列號"].apply(status_for)
        st.dataframe(preview, use_container_width=True, height=300)


# Compact download panel at bottom
st.divider()
st.markdown("<div class='step-title'>Step 4｜下載結果</div>", unsafe_allow_html=True)
st.markdown("<div class='step-caption'>未完成時下載暫存檔；全部處理完成後再下載完成版。</div>", unsafe_allow_html=True)
output_type = st.selectbox(
    "選擇要產生的檔案",
    ["暫存檔（下次可續作，含待確認/失敗/未處理）", "完成版（處理完成後交付）"],
    key="bottom_output_type_v1751",
)

if unfinished > 0:
    st.caption(f"目前未完成：待確認 {len(st.session_state.pending_orders)}｜抓取失敗 {len(st.session_state.failed_orders)}｜手動補卡 {len(st.session_state.manual_orders)}｜未處理 {unprocessed_count}")
else:
    st.caption("目前所有原始列都已處理完成。")

if output_type.startswith("暫存檔"):
    if st.button("產生暫存檔", use_container_width=True, key="btn_build_temp_workbook_bottom_v1751"):
        with st.spinner("正在產生暫存檔..."):
            st.session_state.temp_workbook_bytes = workbook_bytes(setting_bytes, news_df, include_working_states=True)
            st.session_state.download_status_message = "暫存檔已產生。"
    if st.session_state.get("temp_workbook_bytes"):
        st.download_button(
            "下載暫存檔",
            data=st.session_state.temp_workbook_bytes,
            file_name="信用卡新聞月度結果表_暫存版.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key="download_temp_workbook_bottom_v1751",
        )
else:
    if unfinished > 0:
        st.warning("仍有未完成項目，完成版暫不建議產生。請先處理待確認、抓取失敗、手動補卡與未處理新聞。")
    if st.button("產生完成版", use_container_width=True, disabled=unfinished > 0, key="btn_build_final_workbook_bottom_v1751"):
        with st.spinner("正在產生完成版..."):
            st.session_state.final_workbook_bytes = workbook_bytes(setting_bytes, news_df, include_working_states=False)
            st.session_state.download_status_message = "完成版已產生。"
    if st.session_state.get("final_workbook_bytes"):
        st.download_button(
            "下載完成版",
            data=st.session_state.final_workbook_bytes,
            file_name="信用卡新聞月度聲量表_完成版.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            disabled=unfinished > 0,
            key="download_final_workbook_bottom_v1751",
        )

if st.session_state.get("download_status_message"):
    st.caption(st.session_state.download_status_message)

# Auto save progress.
save_progress_to_disk("end_of_run")
